"""
This module has two roles
(I) apply confirmatory factor analyses (CFAs) to trait measures
(II) apply CFAs to resting-state functional connectivity (RSFC) 
(III) apply CFAs and estimate correlation between latent factors of RSFC and trait 
"""

import os
import os.path as op
from pdb import set_trace
from logging import (
    getLogger,
    Formatter,
    FileHandler,
    Filter,
)
import argparse
import pickle
from collections import defaultdict
from warnings import simplefilter
import datetime
from random import sample, shuffle, seed
from itertools import combinations
from functools import reduce
import re
from operator import add
from typing import TypedDict, Literal, Optional, Union
from joblib import Parallel, delayed
from joblib import wrap_non_picklable_objects
from math import comb
from time import time, sleep
from tqdm import tqdm

import numpy as np
from numpy.linalg import inv
from nptyping import NDArray, Shape, Float, Int, Bool
import pandas as pd
from semopy import Model, calc_stats
from semopy.inspector import inspect_matrices
from factor_analyzer.utils import covariance_to_correlation
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import spearmanr
from scipy import linalg as LA
from pingouin import partial_corr, intraclass_corr
from ppca import PPCA
from sklearn.model_selection import GroupKFold
from sklearn.preprocessing import StandardScaler
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
import rpy2
import rpy2.robjects as ro
from rpy2.robjects.packages import importr
import rpy2.robjects.packages as rpackages
from rpy2.robjects import pandas2ri, globalenv

from preprocess import train_test_split_family

HCP_ROOT_DIR = "/home/cezanne/t-haitani/hcp_data"
FAMILY_PATH = '/home/cezanne/t-haitani/hcp_data/RESTRICTED_tomosumi_9_9_2024_3_21_50.csv'
GORDON_DIR = op.join(
    HCP_ROOT_DIR, "derivatives", "Python", "parcellation", "Gordon"
)
SCHAEFER_DIR = op.join(
    HCP_ROOT_DIR, "derivatives", "Python", "parcellation", "Schaefer"
)
ATLAS_DIR_DICT = {
    'Schaefer': SCHAEFER_DIR,
    'Gordon': GORDON_DIR
        }

NEO_FFI_DIR_SCHAEFER = (
    "/home/cezanne/t-haitani/hcp_data/derivatives/Python/parcellation/Schaefer/NEO_FFI"
)
NIH_COGNITION_DIR_SCHAEFER = op.join(SCHAEFER_DIR, "NIH_Cognition")
ASR_DIR_SCHAEFER = op.join(SCHAEFER_DIR, "ASR")
FC_ONLY_DIR_SCHAEFER = op.join(SCHAEFER_DIR, "reliability")
NEO_FFI_DIR_GORDON = (
    "/home/cezanne/t-haitani/hcp_data/derivatives/Python/parcellation/Gordon/NEO_FFI"
)
NIH_COGNITION_DIR_GORDON = op.join(GORDON_DIR, "NIH_Cognition")
ASR_DIR_GORDON = op.join(GORDON_DIR, "ASR")
FC_ONLY_DIR_GORDON = op.join(GORDON_DIR, "reliability")

NIH_COGNITION_DIR_DICT = {
    'Schaefer': NIH_COGNITION_DIR_SCHAEFER,
    'Gordon': NIH_COGNITION_DIR_GORDON
        }
ASR_DIR_DICT = {
    'Schaefer': ASR_DIR_SCHAEFER,
    'Gordon': ASR_DIR_GORDON
        }
NEO_FFI_DIR_DICT = {
    'Schaefer': NEO_FFI_DIR_SCHAEFER,
    'Gordon': NEO_FFI_DIR_GORDON
        }
FC_ONLY_DIR_DICT = {
    'Schaefer': FC_ONLY_DIR_SCHAEFER,
    'Gordon': FC_ONLY_DIR_GORDON
        }

FA_PARAMS_DIR = op.join(HCP_ROOT_DIR, "derivatives", "Python", "fa_result")
SESSIONS = ["rfMRI_REST1_RL", "rfMRI_REST1_LR", "rfMRI_REST2_LR", "rfMRI_REST2_RL"]
COVARIATES_PATH = '/home/cezanne/t-haitani/hcp_data/derivatives/covariates.csv'

#SCALE_NAMES_DICT = {
#        'NIH_Cognition': {
#            'Total': {'Pub': ALL_SUBSCALE_COG_COLUMNS_PUB, 'Original': ALL_SUBSCALE_COG_COLUMNS},
#            'Fluid': {'Pub': FLUID_COGNITION_PUB, 'Original': FLUID_COLUMNS},
#            'Crystal': {'Pub': CRYSTAL_COGNITION_PUB, 'Original': CRYSTAL_COLUMNS}
#            },
#        'ASR': {
#            'All': {'Pub': ASR_ALL_SCALES_PUB, 'Original': ASR_ALL_SCALES},
#            'Internalizing': {'Pub': ASR_INT_SCALES_PUB, 'Original': ASR_INT_SCALES},
#            'Externalizing': {'Pub': ASR_EXT_SCALES_PUB, 'Original': ASR_EXT_SCALES},
#            'Others': {'Pub': ASR_OTHER_SCALES_PUB, 'Original': ASR_OTHER_SCALES}
#            },
#        'NEO_FFI': {
#            'Neuroticism': {'Pub': None, 'Original': []}
#            }
#        }
#
FLUID_COGNITION = ["PicSeq", "CardSort", "Flanker", "ProcSpeed", "ListSort"]
FLUID_COGNITION_PUB = [
    "Picture sequence memory",
    "Dimensional change card sort",
    "Flanker inhibitory control and attention",
    "Pattern comparison",
    "List sorting working memory",
]
CRYSTAL_COGNITION = ["ReadEng", "PicVocab"]
CRYSTAL_COGNITION_PUB = ["Oral reading recognition", "Picture vocabulary"]
FLUID_COLUMNS = [c + "_Unadj" for c in FLUID_COGNITION]
CRYSTAL_COLUMNS = [c + "_Unadj" for c in CRYSTAL_COGNITION]
COMPOSITE_COLUMNS = ["CogFluidComp_Unadj", "CogCrystalComp_Unadj", "CogTotalComp_Unadj"]
ALL_SUBSCALE_COG_COLUMNS = FLUID_COLUMNS + CRYSTAL_COLUMNS
ALL_SUBSCALE_COG_COLUMNS_PUB = FLUID_COGNITION_PUB + CRYSTAL_COGNITION_PUB
ALL_COG_COLUMNS = ALL_SUBSCALE_COG_COLUMNS + COMPOSITE_COLUMNS

FLUID_COLUMNS_AGE_ADJ = [c + "_AgeAdj" for c in FLUID_COGNITION]
CRYSTAL_COLUMNS_AGE_ADJ = [c + "_AgeAdj" for c in CRYSTAL_COGNITION]
ALL_SUBSCALE_COG_COLUMNS_AGE_ADJ = FLUID_COLUMNS_AGE_ADJ + CRYSTAL_COLUMNS_AGE_ADJ
COMPOSITE_COLUMNS_AGE_ADJ = ["CogFluidComp_AgeAdj", "CogCrystalComp_AgeAdj", "CogTotalComp_AgeAdj"]
ALL_COG_COLUMNS_AGE_ADJ = ALL_SUBSCALE_COG_COLUMNS_AGE_ADJ + COMPOSITE_COLUMNS_AGE_ADJ

COG_SCALES_PUB_DICT = {key: value for key, value in zip(ALL_SUBSCALE_COG_COLUMNS, ALL_SUBSCALE_COG_COLUMNS_PUB)}
COG_SCALES_PUB_DICT_REVERSED = {value: key for key, value in COG_SCALES_PUB_DICT.items()}

ASR_BROAD_SCALES = ["ASR_Totp_Raw", "ASR_Intn_Raw", "ASR_Extn_Raw", "ASR_TAO_Sum"]
ASR_BROAD_SCALES_MOD = ["All", "Internalizing", "Externalizing", "Others"]
ASR_SUBSCALES = ["All", "Internalizing", "Externalizing", "Others"]
ASR_INT_SCALES = ["ASR_Anxd_Raw", "ASR_Witd_Raw", "ASR_Soma_Raw"]
ASR_INT_SCALES_PUB = ["Anxious/Depressed", "Withdrawn", "Somatic complaints"]
ASR_EXT_SCALES = ["ASR_Aggr_Raw", "ASR_Rule_Raw", "ASR_Intr_Raw"]
ASR_EXT_SCALES_PUB = ["Aggressive behavior", "Rule Breaking Behavior", "Intrusive"]
ASR_OTHER_SCALES = ["ASR_Thot_Raw", "ASR_Attn_Raw", "ASR_Oth_Raw"]
ASR_OTHER_SCALES_PUB = ["Thought problems", "Attention problems", " Other problems"]
ASR_ALL_SCALES = ASR_INT_SCALES + ASR_EXT_SCALES + ASR_OTHER_SCALES
ASR_ALL_SCALES_PUB = ASR_INT_SCALES_PUB + ASR_EXT_SCALES_PUB + ASR_OTHER_SCALES_PUB

ASR_DICT = {'All': ASR_ALL_SCALES, 'Internalizing': ASR_INT_SCALES, 'Externalizing': ASR_EXT_SCALES, 'Others': ASR_OTHER_SCALES}
ASR_DICT_PUB = {'All': ASR_ALL_SCALES_PUB, 'Internalizing': ASR_INT_SCALES_PUB, 'Externalizing': ASR_EXT_SCALES_PUB, 'Others': ASR_OTHER_SCALES_PUB}

ASR_SCALES_PUB_DICT = {key: value for key, value in zip(ASR_ALL_SCALES, ASR_ALL_SCALES_PUB)}

ModelFA = Literal[
    "model_fc", "model_trait", "model_both", "model_onlyFC", "model_only_trait"
]
MODEL_TRAIT = ["model_fc", "model_trait", "model_both"]
NEO_FFI_SCALES = [
    "Neuroticism",
    "Extraversion",
    "Openness",
    "Agreeableness",
    "Conscientiousness",
]
NEO_FFI_DICT_REVERSED = defaultdict(dict)
for i, scale_name in enumerate(NEO_FFI_SCALES):
    remainder = 0 if i == 4 else i + 1
    NEO_FFI_DICT_REVERSED[scale_name] = {
        value: key for key, value in zip([f'{j:02}' for j in range(1, 61) if j % 5 == remainder], ['Item' + str(i) for i in range(1, 13)])
        }

NEO_FFI_DICT = defaultdict(dict)
for i, scale_name in enumerate(NEO_FFI_SCALES):
    remainder = 0 if i == 4 else i + 1
    NEO_FFI_DICT[scale_name] = [f'NEORAW_{j:02}' for j in range(1, 61) if j % 5 == remainder]

ABB_NEO_FFI_SCALES = ["N", "E", 'O', "A", "C"]
NEO_FFI_ITEMS = [f'NEORAW_{i:02d}' for i in range(1, 61)]

NIH_COGNITION_SCALES = ["Total", "Fluid", "Crystal"]

FIT_INDICES = [
    "DoF",
    "DoF_baseline",
    "chi2",
    "chi2_pvalue",
    "chi2_baseline",
    "CFI",
    "GFI",
    "AGFI",
    "NFI",
    "TLI",
    "RMSEA",
    "AIC",
    "BIC",
    "LogLik",
    "SRMR",
]

LINE_THIN = Side(style="thin")
SIDE = Side(border_style=None)
(BOTTOM_THIN, TOP_THIN_AND_BOTTOM_THIN, TOP_THIN) = (
    Border(bottom=LINE_THIN),
    Border(top=LINE_THIN, bottom=LINE_THIN),
    Border(top=LINE_THIN),
)
NO_BORDER = Border(
    left=SIDE,
    right=SIDE,
    top=SIDE,
    bottom=SIDE,
)
FONT_ARIAL = Font(name="Arial")

class SharedSubscaleName():
    def __init__(self):
        self.ASR_SCALES_PUB_DICT = ASR_SCALES_PUB_DICT
        self.COG_SCALES_PUB_DICT = COG_SCALES_PUB_DICT

TraitType = Literal["cognition", "personality", "mental"]
TRAIT_SCALES = ["NIH_Cognition", "NEO_FFI", "ASR"]
TraitScaleName = Literal[tuple(TRAIT_SCALES)]

SUBSCALE_NAMES = NIH_COGNITION_SCALES + NEO_FFI_SCALES
SubscaleNames = Literal[tuple(SUBSCALE_NAMES)]

ITEM_NAMES = ALL_SUBSCALE_COG_COLUMNS + NEO_FFI_ITEMS
ItemNames = Literal[tuple(ITEM_NAMES)]

lavaan = importr("lavaan")
base = importr("base")


def get_subscale_list(
        trait_scale_name,
        get_all_subscales=False,
        get_reordered_list=False,
        scale_name=None,
        name_type_pub=True
        ):
    """
    function for getting subscale list from trait scale name
    When get_all_subscales is True, scale_name should be specified.
    """
    if trait_scale_name in ["NEO_FFI", 'NEO-FFI']:
        if not get_all_subscales:
            subscale_list = NEO_FFI_SCALES
        else:
            if name_type_pub:
                subscale_list = ['Item' + str(i) for i in range(1, 13)]
            else:
                subscale_list = NEO_FFI_DICT.get(scale_name) 

    elif trait_scale_name in ["NIH_Cognition", 'NIH toolbox']:
        if not get_all_subscales:
            subscale_list = NIH_COGNITION_SCALES
        else:
            if scale_name == 'Total':
                subscale_list = ALL_SUBSCALE_COG_COLUMNS_PUB
            elif scale_name == 'Fluid':
                subscale_list = FLUID_COGNITION_PUB
            elif scale_name == 'Crystal':
                subscale_list = CRYSTAL_COGNITION_PUB

    elif trait_scale_name == "ASR":
        if not get_all_subscales:
            subscale_list = ASR_BROAD_SCALES_MOD
        else:
            if name_type_pub:
                subscale_list = ASR_DICT_PUB.get(scale_name)
            else:
                subscale_list = ASR_DICT.get(scale_name)

    else:
        raise NameError('trait_scale_name is wrong.')

    if get_reordered_list:
        if trait_scale_name in ['NIH_Cognition', 'ASR']:
            subscale_list_alphabet_order = sorted(subscale_list)
            reordered_list = [subscale_list_alphabet_order.index(element) for element in subscale_list]
        elif trait_scale_name == 'NEO_FFI':
            reordered_list = None
    else:
        return subscale_list
    return subscale_list, reordered_list


def generate_ffi_subscale_item_number(
    label, remove_vars_list: Union[list[str], list[None]] = [None]
):
    """generate item numbers of neo-ffi based on input string"""
    ffi_dict = {
        "Neuroticism": 1,
        "Extraversion": 2,
        "Openness": 3,
        "Agreeableness": 4,
        "Conscientiousness": 5,
    }
    number = ffi_dict[label]
    item_number = np.arange(number, 61, 5)
    if remove_vars_list is not None:
        ffi_items = [
            f"NEORAW_{i:02d}" for i in item_number if f"NEORAW_{i:02d}" not in remove_vars_list
        ]
    else:
        ffi_items = [f"NEORAW_{i:02d}" for i in item_number]

    return ffi_items


def generate_NIH_cognition_scales(
    scale_name: str, 
    add_comp: bool, 
    remove_vars_list: Union[list[str], list[None]] = [None],
    age_adj=False
):
    """generate scale names of NIH toolbox cognition"""
    if scale_name == NIH_COGNITION_SCALES[0]:
        cog_scales = ALL_SUBSCALE_COG_COLUMNS if not age_adj else ALL_SUBSCALE_COG_COLUMNS_AGE_ADJ 
        if add_comp:
            add_column_list = ["CogTotalComp_Unadj"] if not age_adj else ['CogTotalComp_AgeAdj']
            cog_scales = cog_scales + add_column_list 
    
    elif scale_name == NIH_COGNITION_SCALES[1]:
        cog_scales = FLUID_COLUMNS if not age_adj else FLUID_COLUMNS_AGE_ADJ
        if add_comp:
            add_column_list = ["CogFluidComp_Unadj"] if not age_adj else ['CogFluidComp_AgeAdj']
            cog_scales = cog_scales + add_column_list
    
    elif scale_name == NIH_COGNITION_SCALES[2]:
        cog_scales = CRYSTAL_COLUMNS if not age_adj else CRYSTAL_COLUMNS_AGE_ADJ
        if add_comp:
            add_column_list = ["CogCrystalComp_Unadj"] if not age_adj else ["CogCrystalComp_Ageadj"]
            cog_scales = cog_scales + add_column_list 
    
    if remove_vars_list is not None:
        cog_scales = [i for i in cog_scales if i not in remove_vars_list]
    
    return cog_scales


def generate_ASR_scales(scale_name, remove_vars_list: Union[list[str], list[None]] = [None]):
    """generate scale names of ASR"""
    if scale_name == "All":
        asr_scales = ASR_ALL_SCALES
    elif scale_name == "Internalizing":
        asr_scales = ASR_INT_SCALES
    elif scale_name == "Externalizing":
        asr_scales = ASR_EXT_SCALES
    elif scale_name == "Others":
        asr_scales = ASR_OTHER_SCALES
    if remove_vars_list is not None:
        asr_scales = [i for i in asr_scales if i not in remove_vars_list]

    return asr_scales


def get_index_of_edges_without_subcortex():
    """get index of edges without those of limbic cortex based on Schaefer parcellation"""
    with open(
        op.join(ATLAS_DIR, "edge_summary_schaefer_400_s2.pkl"),
        "rb",
    ) as f:
        edge_summary_schaefer_s2 = pickle.load(f)
    index_of_edges_without_subcortex = np.array(
        edge_summary_schaefer_s2.query('not node1_net in ["Limbic_tian", "Limbic"]')
        .query('not node2_net == "Limbic"')
        .index
    )
    return index_of_edges_without_subcortex


def read_fc_data(fc_filename, invalid_edge_file=None):
    """function for reading and returning fc data"""
    # read data of FC
    fc_data = np.load(op.join(ATLAS_DIR, fc_filename + ".npy"))
    if invalid_edge_file is not None:
        invalid_edges = np.loadtxt(op.join(ATLAS_DIR, 'reliability', 'invalid_edges', invalid_edge_file)).astype(int)
        fc_data = np.delete(fc_data, invalid_edges, axis=0)
    if fc_data.ndim == 2:
        fc_data = fc_data[:, :, np.newaxis]
    
    return fc_data


def read_trait_data(
        trait_type=None, 
        subjects_filename=None,
        subjects_list=None
        ):
    """
    Read trait data and filter subjects
    """
    if subjects_list is None:
        subjects_list = pd.read_csv(op.join(HCP_ROOT_DIR, 'derivatives', subjects_filename), header=None).loc[:, 0].to_list()
    subjects_list = [str(i) for i in subjects_list]
    if trait_type == 'personality':
        df = read_ffi_data()
        df.columns = [i.replace('NEOFAC_', '') if 'NEOFAC' in i else i for i in df.columns]
    elif trait_type == 'mental':
        df = read_asr_data()
    elif trait_type == 'cognition':
        df = read_cog_data()
        df.columns = [i.replace('Cog', '').replace('Comp_Unadj', '') if 'Cog' in i and 'Unadj' in i else i for i in df.columns]
    if subjects_list:
        df.query('Subject.isin(@subjects_list)', inplace=True)
    return df


def read_ffi_data(remove_missing_subjects_ffi=True):
    """function for reading and returning ffi data"""
    # FFI subscale does not have item-level missing data.
    # Then, this function will conduct only subject-wise deletion.

    # generate item labels to read data
    ffi_composites = ["NEOFAC_" + p for p in ABB_NEO_FFI_SCALES]
    ffi_items = ["NEORAW_" + f"{i:02d}" for i in range(1, 61)]
    ffi_all_scores = ffi_composites + ffi_items

    # read data
    ffi_data = pd.read_csv(
        "/home/cezanne/t-haitani/hcp_data/unrestricted_tomosumi_10_11_2022_1_24_36.csv",
        usecols=["Subject"] + ffi_all_scores,
        dtype={"Subject": "object"},
    )
    for new, old in zip(NEO_FFI_SCALES, ABB_NEO_FFI_SCALES):
        old_name, new_name = f"NEOFAC_{old}", f"NEOFAC_{new}"
        ffi_data.rename(columns={old_name: new_name}, inplace=True)
    # replace strings to integers
    pd.set_option('future.no_silent_downcasting', True)
    ffi_data.replace(
        {"SA": int(4), "A": int(3), "N": int(2), "D": int(1), "SD": int(0)},
        inplace=True,
    )
    #  ffi_data.iloc[:, 1:] = ffi_data.iloc[:, 1:].astype('Int64')

    # make lists of reversed item numbers according to https://doi.org/10.1007/s10519-013-9625-7
    (
        N_reversed_item_numbers,
        E_reversed_item_numbers,
        O_reversed_item_numbers,
        A_reversed_item_numbers,
        C_reversed_item_numbers,
    ) = (
        [1, 16, 31, 46],
        [12, 27, 42, 57],
        [3, 8, 18, 23, 33, 38, 48],
        [9, 14, 24, 29, 39, 44, 54, 59],
        [15, 30, 45, 55],
    )
    ffi_reversed_item_numbers = (
        N_reversed_item_numbers
        + E_reversed_item_numbers
        + O_reversed_item_numbers
        + A_reversed_item_numbers
        + C_reversed_item_numbers
    )
    ffi_reversed_items = ["NEORAW_" + f"{i:02d}" for i in ffi_reversed_item_numbers]
    # reverse scores
    ffi_data[ffi_reversed_items] = ffi_data[ffi_reversed_items].apply(lambda x: 4 - x)
    # calculate total scores
    for subscale_number, p in enumerate(NEO_FFI_SCALES):
        item_numbers = np.arange(subscale_number + 1, 61, 5)
        ffi_data[f"NEO_total_{p}"] = ffi_data[
            ["NEORAW_" + f"{i:02d}" for i in item_numbers]
        ].sum(axis=1)
    # remove subjects with any missing values
    if remove_missing_subjects_ffi:
        ffi_data = ffi_data.loc[~ffi_data.isna().any(axis=1), :]

    return ffi_data


def read_cog_data(cog_missing="remove", age_adj=False):
    """function for reading and returning NIH toolbox cognition data"""
    # read data
    all_columns = ALL_COG_COLUMNS if not age_adj else ALL_COG_COLUMNS_AGE_ADJ
    cog_data = pd.read_csv(
        op.join(HCP_ROOT_DIR, "unrestricted_tomosumi_10_11_2022_1_24_36.csv"),
        usecols=["Subject"] + all_columns,
        dtype={"Subject": "object"},
    )
    
    # set data types of scores
    cog_data.iloc[:, 1:] = cog_data.iloc[:, 1:].astype("float32")
    
    # calculate mean scores considering treatments of missing values
    if not age_adj:
        fluid_columns, crystal_columns, all_subscale_columns = FLUID_COLUMNS, CRYSTAL_COLUMNS, ALL_SUBSCALE_COG_COLUMNS
    else:
        fluid_columns, crystal_columns, all_subscale_columns = FLUID_COLUMNS_AGE_ADJ, CRYSTAL_COLUMNS_AGE_ADJ, ALL_SUBSCALE_COG_COLUMNS_AGE_ADJ

    if cog_missing == "ignore":
        mean_func = np.nanmean
    elif cog_missing == 'remove':
        mean_func = np.mean
    cog_data = cog_data.assign(
        fluid_mean_unadj=lambda x: mean_func(x[fluid_columns], axis=1),
        crystal_mean_unadj=lambda x: mean_func(x[crystal_columns], axis=1),
        total_mean_unadj=lambda x: mean_func(x[all_subscale_columns], axis=1),
    )
    # select columns
    cog_data = cog_data[["Subject"] + all_columns]
    # remove missing values
    if cog_missing == "remove":
        cog_data = cog_data.loc[~cog_data.isna().any(axis=1), :]
    return cog_data


def read_asr_data(remove_missing_subjects_asr=True):
    """function for reading ASR data"""
    ASR_data = pd.read_csv(
        op.join(HCP_ROOT_DIR, "hcp_restricted.csv"),
        usecols=["Subject"] + ASR_BROAD_SCALES + ASR_ALL_SCALES,
        dtype={"Subject": "object"},
    ).rename(
        columns={
            "ASR_Totp_Raw": "All",
            "ASR_Intn_Raw": "Internalizing",
            "ASR_Extn_Raw": "Externalizing",
            "ASR_TAO_Sum": "Others",
        }
    )
    # ASR_data.iloc[:, 1:] = ASR_data.iloc[:, 1:].astype("Int64")
    # remove subjects with any missing values
    if remove_missing_subjects_asr:
        ASR_data = ASR_data.loc[~ASR_data.isna().any(axis=1), :]
    return ASR_data


def read_subject_ids():
    """reading subject IDs"""
    sub_ids = pd.read_csv(
        op.join(HCP_ROOT_DIR, "hcp_restricted.csv"),
        usecols=["Subject"],
        dtype={"Subject": "object"},
    )
    return sub_ids


def generate_drop_prefix(
        remove_vars_list: Optional[list[str]]
        ) -> str:
    """function for generating prefix representing dropped variables"""
    if remove_vars_list is None:
        drop_prefix = ""
    else:
        drop_prefix = "_drop_" + "_".join(remove_vars_list)
    return drop_prefix


def generate_drop_suffix(
    trait_type,
    value_list: list[str]
    ):
    """
    generate suffix of dropped variables from list
    """
    if value_list is not None:
        if trait_type == 'personality':
            value_list = [i.replace('NEORAW_', '') for i in value_list]
        drop_vars_str = '_drop_' + '_'.join(value_list)
    else:
        drop_vars_str = ''
    return drop_vars_str


def generate_drop_suffix_from_trait_scale(trait_type, input_dict):
    """
    generate a string of suffix representing dropped variables
    """
    drop_suffix = ''
    if input_dict is not None:
        for key, value_list in input_dict.get(trait_type).items():
            key_str = '_in_' + key
            drop_vars_str = generate_drop_suffix(trait_type, value_list)
            drop_items_str = drop_vars_str + key_str
            drop_suffix += drop_items_str
    return drop_suffix


def conduct_cfas_to_selected_samples_for_pub(
    model_fit_obj,
    se_robust,
    control,
    fit_indices_list: list[str],
    ffi_ordered=True,
    trait_type_list=["cognition", "mental", "personality"],
    drop_vars_dict: Optional[dict[TraitType, dict[SubscaleNames, list[Optional[ItemNames]]]]] = None,
    save_residuals=False,
    save_table=False,
    filename_suffix: str=None,
    age_adj_cog=False,
    cor_error_list_dict=None,
    subjects_set_for_analysis=None,
    fold_n=None,
    random_seed=None,
    equal_loadings=False,
    **kwargs,
):
    """
    select subjects and then conduct CFAs to trait data
    kwargs should include selection parameters in input_params passed to select_data_for_analysis_mod()
    """
    if subjects_set_for_analysis is None:
        _, _, _, _, subjects_set_for_analysis, _ = select_data_for_analysis_mod(**kwargs)
    subject_n = len(subjects_set_for_analysis)
    
    equal_loading_suffix = 'EqualLoading_' if equal_loadings else ''

    for trait_type in trait_type_list:
        print(f'Processing {trait_type}')
        if trait_type == "personality":
            ordered = ffi_ordered
        else:
            ordered = False
        ordered_suffix = "_polycor" if ordered else ""
        drop_vars_dict_scales = (
            drop_vars_dict.get(trait_type, None) if drop_vars_dict is not None else None
        )
        (
            out_param_pd,
            out_cov_pd,
            out_fit_pd,
            out_omega_pd,
        ) = loop_of_cfa_for_trait_data_for_publication(
            trait_type,
            model_fit_obj,
            se_robust,
            control,
            ordered,
            fit_indices_list,
            subjects_set_for_analysis=subjects_set_for_analysis,
            drop_vars_dict_scales=drop_vars_dict_scales,
            save_residuals=save_residuals,
            age_adj_cog=age_adj_cog,
            cor_error_list_dict=cor_error_list_dict,
            fold_n=fold_n,
            random_seed=random_seed,
            equal_loadings=equal_loadings
        )

        # disply results
        print(trait_type)
        display(out_param_pd)
        display(out_cov_pd)
        display(out_fit_pd)
        display(out_omega_pd)

        # specify directories to save table
        trait_scale_name = select_folder_from_trait(trait_type)
        save_folder = op.join(FA_PARAMS_DIR, trait_scale_name, "tables")
        os.makedirs(save_folder, exist_ok=True)

        if save_table:
            # save as csv file
            out_param_pd, out_fit_pd, out_omega_pd = (
                round(out_param_pd, 3),
                round(out_fit_pd, 3),
                round(out_omega_pd, 3),
            )
            # generate prefix to save file
            drop_suffix = generate_drop_suffix_from_trait_scale(trait_type, drop_vars_dict)
            if random_seed is not None:
                seed_str = f'Seed{random_seed}_'
            else:
                seed_str = ''
            if fold_n is not None:
                fold_str = f'Fold{fold_n}_'
            else:
                fold_str = ''
            prefix = f"{seed_str}{fold_str}{equal_loading_suffix}{trait_scale_name}_N_{subject_n}{ordered_suffix}{drop_suffix}_{model_fit_obj}"
            out_param_pd.to_csv(
                op.join(save_folder, f"{prefix}_loading_table.csv"), float_format="%.3f"
            )
            if control:
                out_cov_pd = round(out_cov_pd, 3)
                out_cov_pd.to_csv(
                    op.join(save_folder, f"{prefix}_cov_table.csv"), float_format="%.3f"
                )
            out_fit_pd.to_csv(
                op.join(save_folder, f"{prefix}_fit_table.csv"), float_format="%.3f"
            )
            out_omega_pd.to_csv(
                op.join(save_folder, f"{prefix}_omega_table.csv"),
                float_format="%.3f",
                index=False,
            )

            # combine csv files of parameters and omegas and save as excel file
            out_param_omega_pd = pd.concat([out_param_pd, out_omega_pd], axis=0)
            sheet_name_param_omega, sheet_name_cov, sheet_name_fit = (
                f"{trait_scale_name}_param_omega",
                f"{trait_scale_name}_cov",
                f"{trait_scale_name}_fit",
            )
            save_path = op.join(save_folder, f"{prefix}_Tables.xlsx")
            writer = pd.ExcelWriter(save_path, engine="xlsxwriter")
            out_param_omega_pd.to_excel(writer, sheet_name=sheet_name_param_omega)
            if control:
                out_cov_pd.to_excel(writer, sheet_name=sheet_name_cov)
            out_fit_pd.to_excel(writer, sheet_name=sheet_name_fit)
            writer.close()

            wb = openpyxl.load_workbook(save_path)

            # edit excel file of parameter estimates and omega coefficients
            ws_param_omega = wb[sheet_name_param_omega]
            # column_widths = []
            ws_param_omega.insert_rows(1, 2)
            ws_param_omega.insert_rows(ws_param_omega.max_row)
            ws_param_omega.merge_cells(
                start_row=1,
                start_column=2,
                end_row=1,
                end_column=ws_param_omega.max_column,
            )
            ws_param_omega["B1"].value = "Model"
            ws_param_omega["A2"].value = "Pattern coefficients"
            ws_param_omega["A3"].value = "Scale"
            rel_cell = "A" + str(ws_param_omega.max_row - 1)
            ws_param_omega[rel_cell].value = "Reliability estimates"

            for i, row in enumerate(ws_param_omega):
                for j, cell in enumerate(row):
                    # if len(column_widths) > j:
                    #     if len(cell) > column_widths[j]:
                    #         column_widths[j] = len(cell)
                    # else:
                    #     column_widths += [len(cell)]
                    ws_param_omega[cell.coordinate].border = NO_BORDER
                    ws_param_omega[cell.coordinate].font = FONT_ARIAL
                    ws_param_omega[cell.coordinate].alignment = Alignment(
                        horizontal="center"
                    )
                cell_A = row[:1][0]
                if not (i == 1) and not (i == ws_param_omega.max_row - 2):
                    cell_A.alignment = Alignment(horizontal="left")
                if i == 0:
                    for cell in row:
                        if cell == "A1":
                            ws_param_omega[cell.coordinate].border = TOP_THIN
                        else:
                            ws_param_omega[
                                cell.coordinate
                            ].border = TOP_THIN_AND_BOTTOM_THIN
                if i == 1:
                    for cell in row:
                        if cell.coordinate != "A2":
                            ws_param_omega[cell.coordinate].border = BOTTOM_THIN
                            cell_column = cell.coordinate[0]
                            cell_row_add_one = int(cell.coordinate[1:]) + 1
                            cell_under_one_row = cell_column + str(cell_row_add_one)
                            ws_param_omega[cell.coordinate].value = ws_param_omega[
                                cell_under_one_row
                            ].value
                            ws_param_omega[cell_under_one_row].value = None
                ws_param_omega["A3"].border = BOTTOM_THIN
                if (2 < i <= ws_param_omega.max_row - 3) or (
                    i == ws_param_omega.max_row - 1
                ):
                    cell_after_B = row[1:]
                    for cell in cell_after_B:
                        if cell.value == None:
                            cell.value = "-"
                        else:
                            cell.number_format = "0.000"
            if control:
                if (i == ws_param_omega.max_row - 3) or (
                        i == ws_param_omega.max_row - 1
                    ):
                        for cell in row:
                            ws_param_omega[cell.coordinate].border = BOTTOM_THIN
            # adjust column width
            # for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            #     ws.column_dimensions[get_column_letter(i)].width = column_width

            wb.save(op.join(save_folder, f"{prefix}_Table{filename_suffix}.xlsx"))


def make_table_rescor_for_pub(
    sample_n: int,
    est_method: str,
    control_vars_list: list[str]=['age', 'gender', 'MeanRMS'],
    drop_vars_dict: dict={
        'NIH_Cognition': {'Fluid': ['Picture sequence memory', 'List sorting working memory']},
        'NEO_FFI': {
            'Openness': ['Item' + str(i) for i in [1, 2, 4, 7, 8]],
            'Agreeableness': ['Item4'],
            'Conscientiousness': ['Item3']
            }
        },
        ):
    """make table of residual correlations for publication"""
    for trait_scale_name in TRAIT_SCALES:
        macro_subscale_list = get_subscale_list(trait_scale_name)
        for subscale_name in macro_subscale_list:
            res_table_dir = op.join(ATLAS_DIR, trait_scale_name, subscale_name, 'residuals', 'model_only_trait')
            file_list = os.listdir(res_table_dir)
            file_list = [i for i in file_list if str(sample_n) in i and est_method in i and not i.startswith('.') and i.endswith('.csv') and all(j in i for j in control_vars_list)]
            item_list = get_subscale_list(trait_scale_name, get_all_subscales=True, scale_name=subscale_name)
            # create index number of items
            for filename in file_list:
                item_list_with_covs = item_list + [i.title() for i in control_vars_list]
                # treat dropped variables
                if 'drop' in filename:
                    drop_vars = drop_vars_dict.get(trait_scale_name).get(subscale_name)
                    item_list_with_covs = [i for i in item_list_with_covs if i not in drop_vars]
                item_index_str_list = [str(i) for i in range(1, len(item_list_with_covs) + 1)] 
                print(filename)
                # read and set dataframe
                df = pd.read_csv(op.join(res_table_dir, filename), index_col=0)
                df.columns = item_index_str_list
                df.index = [number_str + '. ' + item_str for number_str, item_str in zip(item_index_str_list, item_list_with_covs)]
                # df to excel
                filename_without_ext = filename.replace('.csv', '')
                save_path = op.join(res_table_dir, f"{filename_without_ext}.xlsx")
                writer = pd.ExcelWriter(save_path, engine="xlsxwriter")
                df = round(df, 2)
                df.to_excel(writer, sheet_name='table')
                writer.close()
                # load workbook
                wb = openpyxl.load_workbook(save_path)
                # get worksheet for edition
                ws = wb['table']
                # edit worksheet
                for i, row in enumerate(ws):
                    for j, cell in enumerate(row):
                        ws[cell.coordinate].border = NO_BORDER
                        ws[cell.coordinate].font = FONT_ARIAL
                        if j >= 1:
                            ws[cell.coordinate].alignment = Alignment(
                                horizontal="center"
                            )
                            if i >= 1:
                                if j > i:
                                    ws[cell.coordinate].value = ''
                                elif i == j:
                                    ws[cell.coordinate].value = '-'
                                else:
                                    ws[cell.coordinate].value = re.sub(r'0(?=[.])', '', ("%0.2f" % ws[cell.coordinate].value))
                        else:
                            ws[cell.coordinate].alignment = Alignment(
                                horizontal="left"
                        )
                        if i == 0:
                            ws[cell.coordinate].border = TOP_THIN_AND_BOTTOM_THIN
                        if i == len(item_list_with_covs):
                            ws[cell.coordinate].border = BOTTOM_THIN

                # save workbook
                wb.save(op.join(res_table_dir, f"{filename_without_ext}_edited.xlsx"))


def get_workbook(trait_scale_name, file_name):
    """get workbook"""
    file_dir = op.join(ATLAS_DIR, trait_scale_name, 'tables')
    wb = openpyxl.load_workbook(op.join(file_dir, file_name))
    return file_dir, wb


def get_main_worksheet(workbook, trait_scale_name):
    """get main worksheet"""
    main_sheet = workbook[f'{trait_scale_name}_param_omega']
    return main_sheet


def get_nvars_from_trait_scale_name(trait_scale_name):
    if trait_scale_name == 'NIH_Cognition':
        n_vars, n_subscales = 7, 3
    elif trait_scale_name == 'ASR':
        n_vars, n_subscales = 9, 4
    elif trait_scale_name == 'NEO_FFI':
        n_vars, n_subscales = 12, 5
    return n_vars, n_subscales


def replace_file_name(file_name, suffix: str):
    file_name_mod = file_name.replace('.xlsx', '') + f'_{suffix}.xlsx'
    return file_name_mod


def modify_table_cov_fit(
        trait_scale_name: str,
        file_name: str,
        n_cov: int
        ):
    """
    modify table edited in conduct_cfas_to_selected_samples_for_pub to include results of fit indices and covariates
    """
    
    file_dir, wb = get_workbook(trait_scale_name, file_name)

    # read worksheets
    cov_sheet = wb[f'{trait_scale_name}_cov']
    fit_sheet = wb[f'{trait_scale_name}_fit']
    main_sheet = get_main_worksheet(wb, trait_scale_name)
    
    # get a number of max column
    n_col = main_sheet.max_column
    # reformat existing table elements
    # insert a row
    main_sheet.insert_rows(idx=3, amount=1)
    # rename header of items
    item_header_row = 4
    item_header_cell = main_sheet.cell(row=item_header_row, column=1)
    item_header_cell.value, item_header_cell.border = 'Item', NO_BORDER
    # merge cells where a header of pattern coefficients is placed
    pc_spanner_row = 3
    
    n_vars, n_subscales = get_nvars_from_trait_scale_name(trait_scale_name)
    

    def create_spanner(n_row, header: str):
        """create stub header with centering alignment"""
        main_sheet.merge_cells(start_row=n_row, start_column=1, end_row=n_row, end_column=n_subscales + 1)
        target_cell = main_sheet.cell(row=n_row, column=1)
        target_cell.value = header
        target_cell.alignment = Alignment(horizontal="center")
        target_cell.font = FONT_ARIAL 
        return main_sheet


    main_sheet = create_spanner(pc_spanner_row, 'Pattern coefficients')
    # delete string of old item header
    main_sheet.cell(row=pc_spanner_row - 1, column=1).value = ''
    # insert a space before each item of pattern coefficients
    for i in range(item_header_row + 1, item_header_row + 1 + n_vars):
        main_sheet.cell(row=i, column=1).value = f' {main_sheet.cell(row=i, column=1).value}'
    
    # reformat reliability
    rel_spanner_row = item_header_row + n_vars + 1
    main_sheet = create_spanner(rel_spanner_row, 'Reliability estimates')
    
    # specify a number of the row where insertion of fit indices starts
    start_row_fit = 7 + n_vars
    # specify a number of rows where insertion of fit is conducted
    n_row_fit = 3
    # specify a number of rows where insertion of covariates is conducted
    n_row_cov = n_cov

    # create a header of model fits
    main_sheet = create_spanner(start_row_fit, 'Model fits')

    # create a sub-heading of model fits
    fit_index_header_cell = main_sheet.cell(row=start_row_fit + 1, column=1)
    fit_index_header_cell.value, fit_index_header_cell.font = 'Fit index', FONT_ARIAL

    # create a header of effects of covariates
    start_row_cov = start_row_fit + n_row_fit + 2
    main_sheet = create_spanner(start_row_cov, 'Regression coefficients of covariates')
    # create a sub-geading of effects of covariates
    param_header_cell = main_sheet.cell(row=start_row_cov + 1, column=1)
    param_header_cell.value, param_header_cell.font = 'Covariate', FONT_ARIAL

    def copy_value_from_csv(insert_sheet, n_row_start, n_row_insert_values):
        """copy values from csv files"""
        for i in range(1, n_row_insert_values + 1):
            for j in range(1, n_col + 1):
                insert_value = insert_sheet.cell(row=i + 1, column=j).value
                if insert_value in ['age', 'gender']:
                    insert_value = insert_value.title()
                insert_value_formatted = f' {insert_value}' if j == 1 else insert_value
                target_cell = main_sheet.cell(row=n_row_start + i + 1, column=j)
                target_cell.value = insert_value_formatted
                target_cell.font = FONT_ARIAL
                if j > 1:
                    target_cell.alignment = Alignment(horizontal="center")
                if i == n_row_insert_values:
                    target_cell.border = BOTTOM_THIN
        return main_sheet

    # copy cell values from fit sheet
    main_sheet = copy_value_from_csv(fit_sheet, start_row_fit, n_row_fit)
    # copy from cov sheet
    main_sheet = copy_value_from_csv(cov_sheet, start_row_cov, n_row_cov)

    # rename and save file
    filename_mod = replace_file_name(file_name, 'merged') 
    filename_mod = filename_mod.replace('Extraversion', 'E').replace('Openness', 'O').replace('Agreeableness', 'A').replace('Conscientiousness', 'C')
    wb.save(op.join(file_dir, filename_mod))


def add_modified_model_params_in_table(
        trait_scale_name,
        main_file_name,
        modified_model_file_name,
        subscale_name_list: list[str]
        ):
    """
    add values of estimates in modified models in parentheses
    """
    # get workbooks
    file_dir, main_workbook = get_workbook(trait_scale_name, main_file_name)
    _, modified_model_workbook = get_workbook(trait_scale_name, modified_model_file_name)
    # get worksheets
    main_sheet = get_main_worksheet(main_workbook, trait_scale_name)
    modified_model_sheet = get_main_worksheet(modified_model_workbook, trait_scale_name)
    
    def get_column_of_subscales(worksheet, subscale_name):
        """get columns of specified subscales"""
        for j in range(1, worksheet.max_column + 1):
            if worksheet.cell(2, j).value == subscale_name:
                return j
    
    def add_values_in_parentheses(main_sheet, modified_model_sheet, column_n):
        n_vars, _ = get_nvars_from_trait_scale_name(trait_scale_name) 
        # specify row numbers
        row_start_pattern_coefs = 5
        row_start_rels = 5 + n_vars + 1
        row_start_fits = row_start_rels + 3
        row_start_covs = row_start_fits + 5
        row_start_end_dict = {
                'pattern_coef': {'start': 5, 'end': 5 + n_vars},
                'rel': {'start': row_start_rels, 'end': row_start_rels + 1},
                'fits': {'start': row_start_fits, 'end': row_start_fits + 3},
                'covs': {'start': row_start_covs, 'end': row_start_covs + 2}
                }

        def modify_values(row_start_end_dict: dict, n_vars: int):
            for key, row_dict in row_start_end_dict.items():
                for i in range(row_dict.get('start'), row_dict.get('end')):
                    main_sheet_target_cell = main_sheet.cell(row=i, column=column_n)
                    original_value = main_sheet_target_cell.value
                    if not original_value == '-':
                        modified_value = modified_model_sheet.cell(row=i, column=column_n).value
                        if not key == 'covs':
                            if type(modified_value) is float:
                                modified_value = f'{modified_value:.3f}' 
                      #      print(original_value, modified_value)
                            merged_value = f'{original_value:.3f} ({modified_value})'
                        else:
                            merged_value = f'{original_value} ({modified_value})'
                        main_sheet_target_cell.value = merged_value
            return main_sheet

        main_sheet = modify_values(row_start_end_dict, n_vars)
        return main_sheet

    for subscale_name in subscale_name_list:
        column_n = get_column_of_subscales(main_sheet, subscale_name)
        main_sheet = add_values_in_parentheses(main_sheet, modified_model_sheet, column_n)
    
    file_name_mod = replace_file_name(main_file_name, 'add_parentheses')
    main_workbook.save(op.join(file_dir, file_name_mod))


def edit_combined_excel_ws(worksheet, var_type):
    worksheet.merge_cells(
        start_row=1, start_column=2, end_row=1, end_column=worksheet.max_column
    )
    if var_type == "cov":
        worksheet["B1"].value = "Covariate"
    elif var_type == "fit":
        worksheet["B1"].value = "Fit index"
    worksheet["A1"].value = "Model"
    worksheet["A2"] = None
    worksheet.insert_rows(3)
    ASR_position = 3 + len(NIH_COGNITION_SCALES) + 1
    FFI_position = ASR_position + len(ASR_SUBSCALES) + 1
    worksheet.merge_cells(
        start_row=3, start_column=1, end_row=3, end_column=worksheet.max_column
    )
    worksheet.merge_cells(
        start_row=ASR_position,
        start_column=1,
        end_row=ASR_position,
        end_column=worksheet.max_column,
    )
    worksheet.merge_cells(
        start_row=FFI_position,
        start_column=1,
        end_row=FFI_position,
        end_column=worksheet.max_column,
    )
    worksheet["A3"].value = "NIH Toolbox"
    for i, row in enumerate(worksheet):
        for j, cell in enumerate(row):
            worksheet[cell.coordinate].border = NO_BORDER
            worksheet[cell.coordinate].font = FONT_ARIAL
            if i == 0:
                if cell.coordinate[0] == "A":
                    worksheet[cell.coordinate].border = TOP_THIN
                else:
                    worksheet[cell.coordinate].border = TOP_THIN_AND_BOTTOM_THIN
            if i == 1:
                if cell.coordinate[0] != "A":
                    if var_type == "cov":
                        worksheet[cell.coordinate].value = worksheet[
                            cell.coordinate
                        ].value.title()
                    worksheet[cell.coordinate].border = BOTTOM_THIN
            if i in (ASR_position - 2, FFI_position - 2, worksheet.max_row - 1):
                worksheet[cell.coordinate].border = BOTTOM_THIN
            if i + 1 in (ASR_position, FFI_position):
                if not cell.coordinate[0] != "A":
                    worksheet[cell.coordinate] = None
            # center align
            worksheet[cell.coordinate].alignment = Alignment(horizontal="center")
            # left align variables (models)
            if not i + 1 in (1, 3, ASR_position, FFI_position):
                if cell.coordinate[0] == "A":
                    worksheet[cell.coordinate].alignment = Alignment(horizontal="left")
    (
        worksheet["A" + str(ASR_position)].value,
        worksheet["A" + str(FFI_position)].value,
    ) = ("ASR", "NEO-FFI")
    return worksheet


def generate_res_cor_filename(
        scale_name,
        subject_n,
        model_fit_obj,
        control,
        remove_vars_list=None,
        use_lavaan=False
        ):
    """generate filename of residual correlations"""
    control_suffix = "controlling_" + "_".join(control)
    drop_prefix = generate_drop_prefix(remove_vars_list)
    lavaan_suffix = '_lavaan' if use_lavaan else ''
    filename = f"{scale_name}_N_{subject_n}_{model_fit_obj}_{control_suffix}{lavaan_suffix}_{drop_prefix}_res_cor.csv",
    return filename


def r_df_to_pandas_df(r_df):
    """convert r dataframe to pandas dataframe"""
    with (ro.default_converter + pandas2ri.converter).context():
        pd_df = ro.conversion.get_conversion().rpy2py(r_df)
    return pd_df


def loop_of_cfa_for_trait_data_for_publication(
    trait_type,
    model_fit_obj,
    se_robust: bool,
    control: list[str],
    ordered: bool,
    fit_indices_list: list[str],
    subjects_set_for_analysis=None,
    drop_vars_dict_scales: Optional[dict[str, list[str]]] = None,
    save_residuals=False,
    use_lavaan=True,
    residual_type=None,
    age_adj_cog=False,
    cor_error_list_dict=None,
    fold_n=None,
    equal_loadings=False,
    random_seed=None
):
    """function for looping CFAs to trait data"""
    model = "model_only_trait"
    trait_scale_name = select_folder_from_trait(trait_type)
    # if subjects_set_for_analysis is not None:
    #     subject_n = len(subjects_set_for_analysis)
    
    equal_loading_suffix = 'EqualLoadings_' if equal_loadings else ''

    if trait_type == "personality":
        data = read_ffi_data()
        subscales = NEO_FFI_SCALES
        all_subscales = ["Item " + str(i) for i in range(1, 13)]
        all_subscales_pub = all_subscales
    
    elif trait_type == "cognition":
        data = read_cog_data(cog_missing="remove", age_adj=age_adj_cog)
        subscales = NIH_COGNITION_SCALES
        all_subscales = ALL_SUBSCALE_COG_COLUMNS if not age_adj_cog else ALL_SUBSCALE_COG_COLUMNS_AGE_ADJ
        all_subscales_pub = ALL_SUBSCALE_COG_COLUMNS_PUB
    
    elif trait_type == "mental":
        data = read_asr_data()
        subscales = ASR_SUBSCALES
        all_subscales, all_subscales_pub = ASR_ALL_SCALES, ASR_ALL_SCALES_PUB

    nrows_base = len(all_subscales)
    ncols = len(subscales)

    # create array storing parameter outputs
    output_params_array = np.empty(shape=(nrows_base, ncols))
    output_params_array[:] = np.nan

    # create array storing outputs on covariates
    if control is not None:
        output_cov_array = np.empty(
            shape=(len(control), ncols), dtype=object
        )
        output_cov_array[:] = np.nan
        # create array storing asterisks of siginificance
        output_ast_array = np.empty(shape=(len(control), ncols))
        output_ast_array = np.nan

    # create array storing outputs of fit indices
    num_fit = len(fit_indices_list)
    output_fit_array = np.empty(shape=(num_fit, ncols))

    # create storing omega coefficients
    output_omega_array = np.empty(shape=ncols)
    
    # Store arrays of factor scores
    output_fs_dict = {} 
    
    # store validity coefficients of factor scores
    output_valid_list = []

    # prepare control data if necessary
    if control is not None:
        covariates_data = pd.read_csv(COVARIATES_PATH)

    # loop in subscales
    for i, scale_name in enumerate(subscales):
        remove_vars_list = (
            drop_vars_dict_scales.get(scale_name)
            if drop_vars_dict_scales is not None
            else None
        )

        if trait_type == "personality":
            variables = generate_ffi_subscale_item_number(
                scale_name, remove_vars_list=remove_vars_list
            )
            all_scales = generate_ffi_subscale_item_number(scale_name)
            cor_error_list = cor_error_list_dict.get(scale_name) if type(cor_error_list_dict) is dict else None
            model_syntax = generate_model_syntax_ffi(
                model, 
                scale_name,
                ordered,
                control,
                remove_vars_list=remove_vars_list,
                use_lavaan=use_lavaan,
                cor_error_list=cor_error_list,
                equal_loadings=equal_loadings
            )
        elif trait_type == "cognition":
            variables = generate_NIH_cognition_scales(
                scale_name, add_comp=False, remove_vars_list=remove_vars_list, age_adj=age_adj_cog
            )
            model_syntax = generate_model_syntax_cog(
                model,
                scale_name,
                control, 
                remove_vars_list=remove_vars_list,
                age_adj=age_adj_cog,
                equal_loadings=equal_loadings
            )
        elif trait_type == "mental":
            variables = generate_ASR_scales(scale_name, remove_vars_list=remove_vars_list)
            model_syntax = generate_model_syntax_asr(
                model,
                scale_name,
                control,
                remove_vars_list=remove_vars_list,
                equal_loadings=equal_loadings
            )
        print(model_syntax)
        # select data
        data_for_cfa = data[["Subject"] + variables]
        # select subjects
        if subjects_set_for_analysis is not None:
            data_for_cfa = data_for_cfa.query("Subject in @subjects_set_for_analysis")
        if control is not None:
            covariates_data['Subject'] = covariates_data['Subject'].astype(str)
            data_for_cfa = pd.merge(data_for_cfa, covariates_data)
        # model fit
        fit_model = apply_model_fit(
                model_syntax,
                data_for_cfa,
                model_fit_obj,
                use_lavaan=use_lavaan
                )

        # get model fits
        fit_indices = get_model_fits(
            fit_model, 'float32', use_lavaan
        )
        result = get_params_from_model(
            fit_model, 'float32', se_robust, use_lavaan, return_array=False, save_variables=True, ordered=ordered
        )
   
        # calculate SRMR and get residual correlation matrix
        if not use_lavaan:
            srmr, res_cor = calculate_srmr(fit_model, data_for_cfa, out_diff_mat=True)
            fit_indices["SRMR"] = srmr
        # get residual correlations
        else:
            res_cor_df = get_residual_cors(
                    fit_model,
                    return_array=False
                    )
            res_cor_pd = r_df_to_pandas_df(res_cor_df)
        if save_residuals:
            if not use_lavaan:
                var_names = fit_model.names[1][0]
                res_cor_pd = pd.DataFrame(res_cor, columns=var_names, index=var_names)
            save_dir = op.join(
                ATLAS_DIR, trait_scale_name, scale_name, "residuals", model
            )
            os.makedirs(save_dir, exist_ok=True)
            subject_n = data_for_cfa.shape[0]
            filename = generate_res_cor_filename(
                    scale_name,
                    subject_n,
                    model_fit_obj,
                    control,
                    remove_vars_list,
                    use_lavaan=use_lavaan
                    )
 
            res_cor_pd.to_csv(
                op.join(
                    save_dir,
                    filename[0]
                                    )
            )

        # get loadings and store outputs
        if not use_lavaan:
            loadings = result.query('op == "~" & rval== "tf"')["Est. Std"]
        else:
            loadings = result.query('op == "=~"')['est.std']
        if trait_type == "personality":
            variables = [
                all_subscales[i]
                for i, scale in enumerate(all_scales)
                if scale in variables
            ]
        var_index = [all_subscales.index(variables[i]) for i, _ in enumerate(variables)]
        output_params_array[var_index, i] = loadings

        # calculate omega
        output_omega_array[i] = calculate_omega(loadings)

        # get effects of covariates
        if control is not None:
            if not use_lavaan:
                cov_effects = result.query(
                    '(op == "~" & lval== "tf") | (op == "~~" & ((lval == "age" & rval == "gender")))'
                )["Est. Std"]
            else:
                cov_effects = result.query('op == "~"')['est.std']

            # get asterisks of significance
            if not use_lavaan:
                cov_ps = result.query(
                    '(op == "~" & lval== "tf") | (op == "~~" & ((lval == "age" & rval == "gender") | (lval == "gender" & rval == "age")))'
                )["p-value"]
            else:
                cov_ps = result.query('op == "~"')['pvalue']
            cov_asterisks = np.where(
                cov_ps < 0.001,
                "***",
                np.where(cov_ps < 0.01, "**", np.where(cov_ps < 0.05, "*", "")),
            )
            # create array of effects of covariates with asterisks
            cov_effects_with_ast = cov_effects.round(3).astype(str).map(str) + cov_asterisks
            output_cov_array[:, i] = cov_effects_with_ast

        # store fit indices
        if not use_lavaan:
            fit_indices_selected = fit_indices[fit_indices_list]
        else:
            fit_indices_index = [FIT_INDICES.index(i) for i in fit_indices_list]
            fit_indices_selected = fit_indices[fit_indices_index]
        output_fit_array[:, i] = fit_indices_selected

        # get modification indices
        mi_df = get_MIs(fit_model)
        mi_pd_df = r_df_to_pandas_df(mi_df)
        float_columns = mi_pd_df.select_dtypes(include='float64').columns
        mi_pd_df[float_columns] = round(mi_pd_df[float_columns], 3)
        mi_pd_df.sort_values('mi', ascending=False, inplace=True)
        if save_residuals:
            save_dir_mi = op.join(ATLAS_DIR, trait_scale_name, scale_name, 'MIs')
            os.makedirs(save_dir_mi, exist_ok=True)
            filename_mis = filename[0].replace('res_cor', 'mi')
            mi_pd_df.to_csv(op.join(save_dir_mi, filename_mis))

        print(scale_name)
        display(result)
        display(fit_indices)

        # Get factor scores
        fscore_code = "lavPredict(fit_model, method = 'Bartlett')"
        globalenv["fit_model"] = fit_model
        f_score = ro.r(fscore_code)
        output_fs_dict[scale_name] = f_score

        ## calculate validity coefficients
        # item correlations (Rkk)
        r_kk = data_for_cfa.drop('Subject', axis=1).corr()
        s_kf = loadings.to_numpy()[:, np.newaxis]
        w_kf = inv(r_kk).dot(s_kf)
        c_ss = w_kf.T.dot(r_kk).dot(w_kf)
        l_ss = np.sqrt(c_ss[0]) 
        r_fs = s_kf.T.dot(w_kf).dot(1/l_ss)
        output_valid_list.append(r_fs[0])
        print(f'Validity coefficient is {r_fs[0]:.3f}.')
    # create dataframe
    out_param_pd = pd.DataFrame(
        output_params_array, index=all_subscales_pub, columns=subscales
    )
    if control:
        out_cov_pd = pd.DataFrame(
            output_cov_array,
            index=control,
            columns=subscales,
        )
    else:
        out_cov_pd = None
    out_fit_pd = pd.DataFrame(
        output_fit_array, index=fit_indices_list, columns=subscales
    )
    out_omega_pd = pd.DataFrame(
        output_omega_array, index=subscales, columns=["Omega"]
    ).T
    fscore_pd = pd.DataFrame(output_fs_dict)
    sample_n = len(subjects_set_for_analysis)
    
    fscore_filename = f'{equal_loading_suffix}fold_{fold_n}_sampleN_{sample_n}_fscore'
    fscore_validity_filename = f'{equal_loading_suffix}fold_{fold_n}_sampleN_{sample_n}_fscore_validity'
    
    if random_seed is not None:
        seed_str = f'Seed{random_seed}_'
        fscore_filename = seed_str + '_' + fscore_filename
        fscore_validity_filename = seed_str + '_' + fscore_validity_filename

    if control is not None:
        controlling_suffix = '_controlling_' + '_'.join(control) 
        fscore_filename += controlling_suffix
        fscore_validity_filename += controlling_suffix
    fscore_pd.to_csv(op.join(FA_PARAMS_DIR, trait_scale_name, 'tables', f'{fscore_filename}.csv'))
    
    out_valid_pd = pd.DataFrame({'Validity': output_valid_list}, index=subscales)
    out_valid_pd .to_csv(op.join(FA_PARAMS_DIR, trait_scale_name, 'tables', f'{fscore_validity_filename}.csv'))

    # return output_params_array, output_cov_array, output_fit_array
    return out_param_pd, out_cov_pd, out_fit_pd, out_omega_pd


def calculate_omega(loadings: NDArray[Shape["Num_params"], Float]):
    """
    calculate omega reliability estimate from factor loading
    """
    uniqueness = 1 - loadings**2
    denominator = np.sum(loadings) ** 2 + np.sum(uniqueness)
    numerator = np.sum(loadings) ** 2
    omega = numerator / denominator
    return omega


def get_inclusion_runs_from_rms(
    rms_remove_percentage,
    rms_thres,
    fc_data,
):
    """select inclusion runs based on relative root-mean-squared motion (RMS)"""
    # load data of number of volumes above threshold of RMS
    rms_over_summary = np.load(
        op.join(
            "/home/cezanne/t-haitani/hcp_data/derivatives",
            f"RMS_over_{rms_thres}.npy",
        )
    )
    # set thresholds (number of volumes) of removal
    threshold = fc_data.shape[1] * rms_remove_percentage
    # select runs for analysis
    inclusion_runs_rms = rms_over_summary < threshold
    return inclusion_runs_rms


def select_subjects_from_rms(rms_pass_all_or_any, inclusion_runs_rms):
    """select subjects based on get_inclusions_runs_from_rms"""
    if rms_pass_all_or_any == "all":
        subjects_bool_rms = inclusion_runs_rms.all(axis=1)
    elif rms_pass_all_or_any == "any":
        subjects_bool_rms = inclusion_runs_rms.any(axis=1)
    return subjects_bool_rms


def select_subjects_from_fmri_data_existence(fc_data, fmri_run_all_or_any):
    """function for selecting subjects based on existencae op fmri data and rms threshold"""
    # check existence of fmri data
    inclusion_fc_data = np.invert(np.isnan(fc_data))
    # get boolean lists representing fc data existence
    if fmri_run_all_or_any == "all":
        subjects_bool_fc_exist = inclusion_fc_data.all(axis=(0, 2))
    elif fmri_run_all_or_any == "any":
        subjects_bool_fc_exist = inclusion_fc_data.any(axis=(0, 2))
    return subjects_bool_fc_exist


def subset_fc_data_from_subjects(fc_data, subject_list):
    """
    Subset FC data from subject list
    """
    subjects_with_mri_folder = []
    for folder in next(os.walk(op.join(HCP_ROOT_DIR, "data")))[1]:
        subjects_with_mri_folder.append(folder)
    
    subjects_bool_list = [
        i in set(subject_list) for i in subjects_with_mri_folder
    ]
    fc_data = fc_data[:, subjects_bool_list, :]
    return fc_data


def select_subjects_from_fc_data(
    subjects_bool_fc_exist, rms_removal, subjects_bool_rms
):
    """function for selecting subjects from existence of fmri data and rms removal criteria"""
    # get boolean lists of including subjects
    if rms_removal:
        subjects_bool_fc = subjects_bool_fc_exist & subjects_bool_rms
    else:
        subjects_bool_fc = subjects_bool_fc_exist
    # get a list of subjects with fmri data folder
    subjects_with_mri_folder = []
    for folder in next(os.walk(op.join(HCP_ROOT_DIR, "data")))[1]:
        subjects_with_mri_folder.append(folder)
    # generate a set of including subjects
    subjects_set_fmri = {
        subject
        for i, subject in enumerate(subjects_with_mri_folder)
        if subjects_bool_fc[i]
    }
    # return list of subjects and corresponding boolean list based on existence of fmri folders
    return subjects_set_fmri, subjects_with_mri_folder


def select_subjects_from_trait_data(trait_all_or_any, **kwargs):
    """function for selecting subjects based on trait data"""
    # get a list of subjects with ffi data
    subjects_with_ffi_data = set(kwargs["ffi_data"]["Subject"])
    subjects_with_cog_data = set(kwargs["cog_data"]["Subject"])
    # get a list of subjects with complete data (no missing values)
    if trait_all_or_any == "all":
        subjects_set_trait = set(subjects_with_ffi_data) & set(subjects_with_cog_data)
    # get a list of subjects with any data (including missing values)
    elif trait_all_or_any == "any":
        subjects_set_trait = set(subjects_with_ffi_data) | set(subjects_with_cog_data)
    # return a subject set
    return subjects_set_trait


def select_data_for_analysis_mod(
    **kwargs,
):
    """
    function for selecting subjects for analysis
    arguments necessary for selecting subjects are following:
    fc_filename, cog_missing, rms_removel, rms_remove_percentage, rms_thres, rms_pass_all_or_any, fmri_run_all_or_any, trait_all_or_any
    """
    # read data
    fc_data = read_fc_data(kwargs["fc_filename"], kwargs['invalid_edge_file'])
    ffi_data = read_ffi_data()
    cog_data = read_cog_data(kwargs["cog_missing"])
    asr_data = read_asr_data()
    # select runs which passed rms threshold if rms_removal is True
    if kwargs["rms_removal"]:
        inclusion_runs_rms = get_inclusion_runs_from_rms(
            kwargs["rms_remove_percentage"], kwargs["rms_thres"], fc_data
        )
        subjects_bool_rms = select_subjects_from_rms(
            kwargs["rms_pass_all_or_any"], inclusion_runs_rms
        )
    # select subjects from fmri data existence
    subjects_bool_fc_exist = select_subjects_from_fmri_data_existence(
        fc_data, kwargs["fmri_run_all_or_any"]
    )
    # select subjects for analysis from fc data considering existence of data and rms thresholds
    subjects_set_fmri, subjects_with_mri_folder = select_subjects_from_fc_data(
        subjects_bool_fc_exist, kwargs["rms_removal"], subjects_bool_rms
    )
    # select subjects for analysis from trait data
    subjects_set_trait = select_subjects_from_trait_data(
        kwargs["trait_all_or_any"],
        ffi_data=ffi_data,
        cog_data=cog_data,
        asr_data=asr_data,
    )
    # select subjects from fc, trait, and RMS data
    subjects_set_for_analysis = subjects_set_fmri & subjects_set_trait
    # Subset subjects with counter-balanced rfMRI data
    if select_cb:
        subjects_cb = np.loadtxt('/home/cezanne/t-haitani/hcp_data/derivatives/subjects_cb.txt', dtype=str)
        subjects_set_for_analysis = subjects_set_for_analysis & set(subjects_cb)
    return (
        fc_data,
        ffi_data,
        cog_data,
        asr_data,
        subjects_set_for_analysis,
        subjects_with_mri_folder,
    )


def transform_and_subset_fc_data(
    fc_data, subjects_bool_list_for_fc_analysis, fisher_z, exclude_subcortex
):
    """transforming and subsetting fc data"""
    fc_data = fc_data[:, subjects_bool_list_for_fc_analysis, :]
    if fisher_z:
        fc_data = np.arctanh(fc_data)
    if exclude_subcortex:
        index_of_edges_without_subcortex = get_index_of_edges_without_subcortex()
        fc_data = fc_data[index_of_edges_without_subcortex, :, :]
    return fc_data


def subsetting_ffi_data(ffi_data, scale_name, subjects_set_for_analysis):
    """subset subscale of ffi and select subjects"""
    ffi_data["Subject"] = ffi_data["Subject"].astype(str)
    ffi_data = ffi_data.query("Subject in @subjects_set_for_analysis")
    total_scale_name = "NEOFAC_" + scale_name
    ffi_data = ffi_data[
        ["Subject"]
        + generate_ffi_subscale_item_number(scale_name)
        + [total_scale_name]
    ]
    ffi_data.rename(columns={total_scale_name: 'Total'}, inplace=True)
    return ffi_data


def subsetting_cog_data(cog_data, scale_name, subjects_set_for_analysis):
    """subset scales of NIH toolbox cognition data and select subjects"""
    # select subjects
    cog_data = cog_data.query("Subject in @subjects_set_for_analysis")
    # select columns
    cog_scales = generate_NIH_cognition_scales(scale_name, add_comp=True)
    cog_data = cog_data[["Subject"] + cog_scales]
    return cog_data


def subsetting_asr_data(asr_data, subjects_set_for_analysis):
    """subset subjects of ASR data"""
    asr_data = asr_data.query("Subject in @subjects_set_for_analysis")
    return asr_data


def generate_suffix(
        edge_total, 
        subjects_set_for_analysis, 
        kfold_i
        ):
    """generate suffix for output files"""
    # gsr
    gsr_type = "gs" if "gs" in fc_filename else "nogs"
    
    # control
    control_suffix = (
        "_controlling_" + "_".join(control)
        if control is not None
        else ""
    )
    control_before_suffix = (
        "_controllingBefore_" + "_".join(control_before)
        if control_before is not None
        else ""
    )

    # sample split
    split_suffix = "_split" if split_ratio is not None else ""
    split_suffix = '_split_half_family' if split_half_family else ''
    seed_suffix = f'_seed{random_seed}' if split_half_family else ''

    # correlation type (z or r)
    cor = "_z" if fisher_z else "_r"
    
    # ordered
    ordered_suffix = "_ordered" if ordered else ""
    
    # sample size
    sample_n = len(subjects_set_for_analysis)
    
    # covariance (correlation) of covariate
    cov_cor_suffix = "_CovCor" if cov_cor else ""
    # CU
    cu_suffix = '_CU' if CU else ''
    # OM
    om_suffix = '_OM' if OM else ''
    # phase encoding direction
    pe_suffix = "_PE" if phase_encoding else ""
    # m -1 method factors
    m1_suffix = 'mMinus1' if m_minus_1 else ''
    # day correlation suffix
    day_cor_suffix = "_DayCor" if day_cor else ""
    # day order suffix
    day_order_suffix = '_OrderInDay' if order_in_day else ''
    # multistate single trait suffix
    msst_suffix = '_MSST' if multistate_single_trait else ''
    st_suffix = '_SingleTrait' if single_trait else ''
    # Bifactor suffix
    bifactor_suffix = '_Bifactor' if bi_factor else ''
    # different loadings
    diff_load_suffix = 'DL' if diff_load else ''
    # eqaulity constraints on trait scale
    trait_equal_loadings_suffix = '_TEL' if equal_loadings else ''
    # std results suffix
    output_std_suffix = '_OutStd' if std_lv else ''

    # method marker suffix
    method_marker_suffix = '_addMarker' if add_method_marker else ''
    # method correlation suffix
    method_cor_suffix = '_methodCor' if method_cor else ''

    # lavaan or semopy
    lavaan_suffix = "_lavaan" if use_lavaan else ""
    
    # day or session suffix
    if fc_unit == 'session':
        fc_unit_suffix = '_session'
    elif fc_unit == 'day':
        fc_unit_suffix = '_day'
    elif fc_unit == 'full':
        fc_unit_suffix = '_full'
    
    # suffix on standardization
    std_suffix = "_std" if cfa_standardize else ""
    
    # suffix on removed varaibles
    drop_suffix = generate_drop_suffix(trait_type, remove_vars_list)
    
    # suffix on fixed loadings
    fix_load_suffix = '_FixedLoad' if fix_loadings_to_one else ''

    # Suffix on inter-factor covariance
    fix_cov_suffix = '_Cov0' if fix_cov_to_zero else ''
    
    # suffix on latent variable regression
    fc_to_trait_suffix = '_FCToTrait' if fc_to_trait else ''

    # time stamp
    time_memmap = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # SLURM array ID
    array_suffix = f'_{array_id}' if array_id is not None else ''
    
    # Mean structure
    mean_str_suffix = '_MeanStr' if mean_structure else ''
    
    # Select counter-balanced subjects suffix
    select_cb_suffix = '_SelectCB' if select_cb else ''

    # generate suffix
    suffix = f"Trait_{trait_type}_Scale_{scale_name}_sampleN_{sample_n}_Fold_{kfold_i}_{gsr_type}_edgeN_{edge_total}_Est_{model_fit_obj}{ordered_suffix}{fix_cov_suffix}{cor}{control_suffix}{control_before_suffix}{cu_suffix}{om_suffix}{m1_suffix}{pe_suffix}{day_cor_suffix}{day_order_suffix}{st_suffix}{msst_suffix}{bifactor_suffix}{diff_load_suffix}{trait_equal_loadings_suffix}{fc_to_trait_suffix}{output_std_suffix}{lavaan_suffix}{method_marker_suffix}{method_cor_suffix}{std_suffix}{mean_str_suffix}{fc_unit_suffix}{drop_suffix}{split_suffix}{select_cb_suffix}{seed_suffix}_{time_memmap}{array_suffix}.dat"

    models = "Model_" + "_".join([i.replace("model_", "") for i in model_type])
    return suffix, models, gsr_type


def set_logger(suffix, models, trait_type, cor_type):
    """function for logging warnings and/or errors"""
    logger = getLogger()
    logger.propagate = False

    if trait_type == "personality":
        folder = NEO_FFI_DIR
    elif trait_type == "cognition":
        folder = NIH_COGNITION_DIR
    elif trait_type == "mental":

        folder = ASR_DIR
    elif trait_type is None:
        folder = FC_ONLY_DIR
    
    parent_folder = op.join(folder, "log")
    os.makedirs(parent_folder, exist_ok=True)
    file_handler = FileHandler(
        op.join(
            parent_folder,
            f"{cor_type}_{models}_{suffix}.log",
        )
    )
    return file_handler, logger


def prepare_covariates(save_file=False):
    """
    Prepare dataframe including covariates inclduing, age, gender, and acquisition time of rs-fMRI
    """
    covariates_data = pd.read_csv(
        op.join(HCP_ROOT_DIR, "unrestricted_tomosumi_10_11_2022_1_24_36.csv"),
        usecols=["Subject", "Gender"],
        dtype="object",
    ).assign(gender=lambda x: x.Gender.replace(["F", "M"], [0, 1]))
    cov_data_restricted = pd.read_csv(
        op.join(HCP_ROOT_DIR, "hcp_restricted.csv"),
        usecols=["Subject", "Age_in_Yrs"],
        dtype={"Subject": object, "Age_in_Yrs": int},
    ).rename(columns={"Age_in_Yrs": "age"})
    covariates_data = pd.merge(covariates_data, cov_data_restricted)
    
    session_info_dir = op.join(HCP_ROOT_DIR, 'sessionSummaryCSV_1200Release')
    subject_list = [i.replace('.csv', '') for i in os.listdir(session_info_dir) if not i.startswith('.~')]
    time_df = pd.DataFrame(columns=['Subject', 'day1_time', 'day2_time', 'delta_hour'])
    
    for i, subject in tqdm(enumerate(subject_list)):
        #print(subject)
        subject_df = pd.read_csv(op.join(session_info_dir, f'{subject}.csv'))
        subject_df.columns = [c.replace(' ', '_') for c in subject_df.columns]
        day1_time_df = subject_df.query('Scan_Description in ["rfMRI_REST1_RL", "rfMRI_REST1_LR"] & Scan_Type == "rfMRI"')['Acquisition_Time']
        day2_time_df = subject_df.query('Scan_Description in ["rfMRI_REST2_RL", "rfMRI_REST2_LR"] & Scan_Type == "rfMRI"')['Acquisition_Time']
        
        if len(day1_time_df) > 0:
            day1_mean_time = pd.to_datetime(day1_time_df, format='%H:%M:%S').mean()
            day1_mean_time_df = day1_mean_time.strftime('%H:%M:%S')
        else:
            day1_mean_time, day1_mean_time_df = None, None

        if len(day2_time_df) > 0:
            day2_mean_time = pd.to_datetime(day2_time_df, format='%H:%M:%S').mean()
            day2_mean_time_df = day2_mean_time.strftime('%H:%M:%S')
        else:
            day2_mean_time, day2_mean_time_df = None, None

        if day1_mean_time is not None and day2_mean_time is not None:
            delta_hour = (day1_mean_time - day2_mean_time).total_seconds() / 3600
        else:
            delta_time = None
        time_df.loc[i, ['Subject', 'Day1Time', 'Day2Time', 'delta_hour']] = subject, day1_mean_time_df, day2_mean_time_df, delta_hour 


    for column in time_df.columns:
        if 'Time' in column:
            # https://stackoverflow.com/questions/24588437/convert-date-to-float-for-linear-regression-on-pandas-data-frame
            time_df[column + 'Float'] = (pd.to_datetime(time_df[column]) - pd.to_datetime(time_df[column]).min()) / np.timedelta64(1,'D')
    covariates_df = pd.merge(covariates_data, time_df, on='Subject')
    
    rms_array = np.empty(shape=(len(subject_list), 4))
    rms_array[:] = np.nan
    for i, subject in tqdm(enumerate(subject_list)):
        for j, session in enumerate(SESSIONS):
            target_file = op.join(HCP_ROOT_DIR, 'data', subject, session, 'Movement_RelativeRMS_mean.txt')
            if op.isfile(target_file):
                rms = np.loadtxt(target_file)
                rms_array[i, j] = rms
    rms_means = np.nanmean(rms_array, axis=1)
    rms_df = pd.DataFrame({'Subject': subject_list, 'MeanRMS': rms_means})
    covariates_df = pd.merge(covariates_df, rms_df, on='Subject')
    if save_file:
        covariates_df.to_csv(COVARIATES_PATH, index=False)


def get_covariates(
    control, control_before, subjects_bool_list_for_fc_analysis, subjects_set_for_analysis
):
    """get covariates"""
    if control is not None or control_before is not None:
        covariates_df = pd.read_csv(COVARIATES_PATH, dtype={'Subject': str})
        covariates_df = covariates_df.query(
            "Subject in @subjects_set_for_analysis"
        )
    else:
        covariates_df = None
    return covariates_df


def select_folder_from_trait(trait_type):
    """function for selecting a folder saving outputs"""
    if trait_type == "personality":
        folder = "NEO_FFI"
    elif trait_type == "cognition":
        folder = "NIH_Cognition"
    elif trait_type == "mental":
        folder = "ASR"
    elif trait_type is None:
        folder = "reliability"
    # else:
    #      raise NameError('trait_type should be "personality" or "cognition".')
    return folder


def get_latest_file(filename_list):
    """
    Get latest file from list
    """
    filename_list = [i for i in filename_list if '.dat' in i]
    filename_list.sort(
        key=lambda x: datetime.datetime.strptime(
            re.findall(
                "(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).dat", x
                )[0],
            "%Y-%m-%d %H:%M:%S"
        )
    )
    return filename_list[-1]


### functions for generating memmap files storing outputs
def generate_cor_memmap(num_iter, suffix, models, **kwargs):
    """generate mammap file storing outputs of correlation estimates"""
    output_cor_name = f"{cor_type}_{models}_{suffix}"

    cor_out_shape = (num_iter, len(model_type))
    if trait_type:
        folder = select_folder_from_trait(trait_type)
        cor_folder = op.join(ATLAS_DIR, folder, scale_name, "correlation")
        os.makedirs(cor_folder, exist_ok=True)
        cor_memmap = generate_or_wait_memmap(
            edge_start,
            cor_folder,
            cor_out_shape,
            output_cor_name,
            dtype_memmap
            )

        return cor_memmap
    else:
        pass


def generate_param_memmap(
    model, param_num, num_iter, suffix, **kwargs
) -> NDArray[Shape["Num_iter, Num_param, Num_cols"], Float]:
    """
    generate mammap file storing outputs of unstandardied and standardized parameter esatimates and
    standard errors of unstandardized parameter estimates
    """
    # adjust output file names and number of parameters
    output_param_name = f"params_{model}_{suffix}".replace("model", "Model")
    if kwargs["use_lavaan"] is False:
        param_ncols = 3
    else:
        param_ncols = 5
    param_out_shape = (num_iter, param_num, param_ncols)

    # select folder for saving file based on trait
    if not model == "model_onlyFC":
        folder = select_folder_from_trait(trait_type)
        param_folder = op.join(ATLAS_DIR, folder, scale_name, "parameters")
    else:
        param_folder = op.join(ATLAS_DIR, "reliability", "parameters")

    os.makedirs(param_folder, exist_ok=True)
    # generate memmap file
    param_memmap = generate_or_wait_memmap(
        edge_start,
        param_folder,
        param_out_shape,
        output_param_name,
        dtype_memmap
        )
    return param_memmap


def generate_or_wait_memmap(
        edge_start,
        folder,
        memmap_shape,
        filename,
        dtype_memmap
        ):
    """
    Generate or wait until generating memory map
    """
    memmap_path = op.join(folder, filename)
    # generate memmap file
#    if edge_start is None:
    memmap_file = np.memmap(
        filename=memmap_path,
        shape=memmap_shape,
        mode="w+",
        dtype=dtype_memmap,
    )
    print(f'Created memmap file : {filename} with {memmap_shape}')
#    else:
#        # Get latest memmap file
#        memmap_file = np.memmap(
#            filename=memmap_path,
#            shape=memmap_shape,
#            mode="w+",
#            dtype=dtype_memmap,
#        )
#        print(f'Created memmap file: {filename}')
    return memmap_file


def generate_fscore_memmap(
    model, sample_n, num_iter, suffix, add_method_marker, **kwargs
) -> NDArray[Shape["Num_iter, Sample_n, 2"], Float]:
    """
    function for generating memmap file storing factor scores
    """
    memmap_name = f"factor_scores_{model}_{suffix}".replace("model", "Model")
    if trait_type:
        shape_dim2 = 2 if not add_method_marker else 3
    else:
        shape_dim2 = 1 if not add_method_marker else 3
    
    if multistate_single_trait or bi_factor:
        shape_dim2 += 2
        if OM:
            if phase_encoding ^ order_in_day:
                if not m_minus_1:
                    shape_dim2 += 2
                else:
                    shape_dim2 += 1
    fscore_memmap_shape = (num_iter, sample_n, shape_dim2)
    if not model == "model_onlyFC":
        folder = select_folder_from_trait(trait_type)
        memmap_folder = op.join(ATLAS_DIR, folder, scale_name, "factor_scores")
    else:
        memmap_folder = op.join(ATLAS_DIR, "reliability", "factor_scores")
    os.makedirs(memmap_folder, exist_ok=True)
    fscore_memmap = generate_or_wait_memmap(
        edge_start,
        memmap_folder,
        fscore_memmap_shape,
        memmap_name,
        dtype_memmap
        )

    return fscore_memmap


def generate_resid_memmap(
    model, var_num, num_iter, suffix, **kwargs
) -> NDArray[Shape["Num_iter, Num_var, Num_var"], Float]:
    """
    generate mammap file storing outputs of standardized residuals
    """
    # adjust output file names and number of parameters
    output_resid_name = f"std_residuals_{model}_{suffix}".replace("model", "Model")
    res_out_shape = (num_iter, var_num, var_num)

    # select folder for saving file based on trait
    if not model == "model_onlyFC":
        folder = select_folder_from_trait(trait_type)
        resid_folder = op.join(ATLAS_DIR, folder, scale_name, "residuals")
    else:
        resid_folder = op.join(ATLAS_DIR, "reliability", "residuals")
    os.makedirs(resid_folder, exist_ok=True)
    # generate memmap file
    
    resid_memmap = generate_or_wait_memmap(
        edge_start,
        resid_folder,
        res_out_shape,
        output_resid_name,
        dtype_memmap
        )

    return resid_memmap


def generate_model_implied_vcov_memmap(
    model, var_num, num_iter, suffix, **kwargs
) -> NDArray[Shape["Num_iter, Num_var, Num_var"], Float]:
    """
    generate mammap file storing outputs of model implied variance-covariance matrix
    """
    # adjust output file names and number of parameters
    output_name = f"model_implied_vcov_{model}_{suffix}".replace("model", "Model")
    out_shape = (num_iter, var_num, var_num)

    # select folder for saving file based on trait
    if not model == "model_onlyFC":
        folder_ = select_folder_from_trait(trait_type)
        folder = op.join(ATLAS_DIR, folder_, scale_name, "model_vcov")
    else:
        folder = op.join(ATLAS_DIR, "reliability", "model_vcov")
    os.makedirs(folder, exist_ok=True)
    # generate memmap file
    
    vcov_memmap = generate_or_wait_memmap(
        edge_start,
        folder,
        out_shape,
        output_name,
        dtype_memmap
        )

    return vcov_memmap


def generate_fit_indices_memmap(
    num_iter, suffix, **kwargs
) -> NDArray[Shape["Num_iter, 15, Model_num"], Float]:
    """generate memmap file storing outputs of fit indices"""
    model_fa = get_factor_models(model_type)
    model_num = len(model_fa)
    models_fa = "Model_" + "_".join([i.replace("model_", "") for i in model_fa])
    output_fit_name = f"fit_indices_{models_fa}_{suffix}"
    fit_out_shape = (num_iter, 15, model_num)
    if not "model_onlyFC" in model_fa:
        folder = select_folder_from_trait(trait_type)
        memmap_folder = op.join(ATLAS_DIR, folder, scale_name, "fit_indices")
    else:
        memmap_folder = op.join(ATLAS_DIR, "reliability", "fit_indices")

    os.makedirs(memmap_folder, exist_ok=True)
    fit_memmap = generate_or_wait_memmap(
        edge_start,
        memmap_folder,
        fit_out_shape,
        output_fit_name,
        dtype_memmap
        )

    return fit_memmap


def get_factor_models(model_type):
    """function for getting number of factor models in model_type"""
    model_fa = [m for m in model_type if "model" in m]
    return model_fa


def settings_and_preprocesses(
    subjects_set_for_analysis,
    subjects_with_mri_folder,
    fc_data,
    ffi_data,
    cog_data,
    asr_data,
    kfold_i,
    **kwargs,
):
    """
    function for settings and preprocesses, including subject selection, edge selection,
    generating memmap files storing results, setting logger, and preparing covariate dataset
    """
    subjects_bool_list_for_fc_analysis = [
        i in subjects_set_for_analysis for i in subjects_with_mri_folder
    ]
    
    # filter warnings
    simplefilter(action="ignore", category=FutureWarning)
    # transform and subset fc data
    fc_data = transform_and_subset_fc_data(
        fc_data, subjects_bool_list_for_fc_analysis, fisher_z, exclude_subcortex
    )

    # subset trait data
    # reset_index() is necessary for concatenation with fc data
    if trait_type == "personality":
        trait_data = subsetting_ffi_data(
            ffi_data, scale_name, subjects_set_for_analysis
        ).reset_index(drop=True)
    elif trait_type == "cognition":
        trait_data = subsetting_cog_data(
            cog_data, scale_name, subjects_set_for_analysis
        ).reset_index(drop=True)
    elif trait_type == "mental":
        trait_data = subsetting_asr_data(
            asr_data, subjects_set_for_analysis
        ).reset_index(drop=True)
    # remove specified variables
    if remove_vars_list is not None:
        trait_data = trait_data.drop(remove_vars_list, axis=1)
        if trait_type == 'personality':
            trait_data['Total'] = trait_data.drop(['Subject', 'Total'], axis=1).sum(axis=1)
        elif trait_type == 'cognition':
            total_scale_name = trait_data.columns[-1]
            trait_data[total_scale_name] = trait_data.drop(['Subject', total_scale_name], axis=1).mean(axis=1)
    # should clarify why this process is conducted (may be processing for rfmri data only)
    elif trait_type is None:
        trait_data = (
            read_subject_ids()
            .query("Subject in @subjects_set_for_analysis")
            .reset_index(drop=True)
        )
    sample_n = len(subjects_set_for_analysis)
    print(sample_n)
    # generate strings for saving outputs
    if save_data:
        # prefix
        suffix, models, gsr_type = generate_suffix(
            edge_total, subjects_set_for_analysis, kfold_i
        )
        # set logger
        file_handler, logger = set_logger(suffix, models, trait_type, cor_type)
        # generate memmap files
        if (not "model_onlyFC" in model_type) or (trait_type is not None) and edge_start is None:
            cor_memmap = generate_cor_memmap(num_iter, suffix, models, **input_params)
        if any(i in ['model_onlyFC'] + MODEL_TRAIT for i in model_type):
            fit_memmap = generate_fit_indices_memmap(num_iter, suffix, **input_params)
        else:
            fit_memmap = None
    else:
        suffix, models, gsr_type = None, None, None
    # get covariate data
    covariates_data = get_covariates(
        control, control_before, subjects_bool_list_for_fc_analysis, subjects_set_for_analysis
    )
    if control_before:
        covs = covariates_data[control_before].to_numpy()
        ## conduct confound regression
        # FC
        thetas = inv(covs.T.dot(covs)).dot(covs.T).dot(fc_data)
        preds_array = np.empty(shape=fc_data.shape)
        preds_array[:] = np.nan
        for i in range(fc_data.shape[2]):
            preds = covs.dot(thetas[..., i])
            preds_array[..., i] = preds.T
        fc_data -= preds_array
        # trait
    data_dict = {
        "fc_data": fc_data,
        "trait_data": trait_data,
        "covariates_data": covariates_data,
    }
    if save_data:
        if trait_type is not None:
            memmap_dict = {"cor_memmap": cor_memmap, "fit_memmap": fit_memmap}
        else:
            memmap_dict = {"fit_memmap": fit_memmap}
        logger_dict = {"file_handler": file_handler, "logger": logger}
        string_dict = {"suffix": suffix, "gsr_type": gsr_type}
    else:
        memmap_dict, logger_dict, string_dict = {}, {}, {}

    return (
        data_dict,
        memmap_dict,
        logger_dict,
        string_dict,
        num_iter,
        sample_n,
        suffix,
    )


def pca(data, pca_z_scaling=True, dims_rescaled_data=1):
    """
    returns: data transformed in 2 dims/columns + regenerated original data
    pass in: data as 2D NumPy array
    """
    if pca_z_scaling:
        scaler = StandardScaler()
        data = scaler.fit_transform(data)

    #   m, n = data.shape
    # mean center the data
    data -= data.mean(axis=0)
    # calculate the covariance matrix
    r = np.cov(data, rowvar=False)
    # calculate eigenvectors & eigenvalues of the covariance matrix
    # use 'eigh' rather than 'eig' since R is symmetric,
    # the performance gain is substantial
    evals, evecs = LA.eigh(r)
    # sort eigenvalue in decreasing order
    idx = np.argsort(evals)[::-1]
    evecs = evecs[:, idx]
    # sort eigenvectors according to same index
    evals = evals[idx]
    # select the first n eigenvectors (n is desired dimension
    # of rescaled data array, or dims_rescaled_data)
    evecs = evecs[:, :dims_rescaled_data]
    # carry out the transformation on the data using eigenvectors
    # and return the re-scaled data, eigenvalues, and eigenvectors
    return np.squeeze(np.dot(evecs.T, data.T)), evecs


def generate_model_syntax_control(control, model, cov_cor=True):
    """function for generating model syntax for controls"""
    if model in ["model_only_trait", 'model_onlyFC']:
        factor = 'tf' if 'trait' in model else 'ff'
        control_syntax = "\n" + f"{factor} ~ " + " + ".join([i for i in control if not 'Time' in i])
    else:
        control_regress_fc = ' + '.join([i for i in control if not 'Time' in i])
        control_regress_trait = ' + '.join([i for i in control if not 'Time' in i]) 
        control_syntax = (
            "\n" + "ff ~ " + control_regress_fc + "\n" + "tf ~ " + control_regress_trait
        )

    if 'age' in control and 'gender' in control and 'MeanRMS' in control:
        if cov_cor:
            cov_syntax = "age ~~ gender\ngender ~~ MeanRMS\nage ~~ MeanRMS"
            control_syntax += "\n" + cov_syntax
    
    if 'Day1TimeFloat' in control and 'Day2TimeFloat' in control:
        if cov_cor:
        #    cov_syntax_time =  "Day1TimeFloat ~~ Day2TimeFloat"
            cov_syntax_time = ''
            control_syntax += "\n" + cov_syntax_time
        #if fc_unit == 'session':
        #    cov_syntax_time_on_ind = '\ns1 ~ Day1TimeFloat\ns2 ~ Day1TimeFloat\ns3 ~ Day2TimeFloat\ns4 ~ Day2TimeFloat'
        #    control_syntax += cov_syntax_time_on_ind
        cov_syntax_time_on_occasion = '\no1 ~ Day1TimeFloat\no2 ~ Day2TimeFloat'
        control_syntax += cov_syntax_time_on_occasion
    return control_syntax


def generate_model_syntax_fc(
    model,
    control,
    phase_encoding=True,
    day_cor=True,
    fc_unit='session',
    order_in_day=False,
    add_method_marker=False,
    multistate_single_trait=False,
    bi_factor=False
):
    """
    function for generating model syntax on FC
    """
    if fc_unit == 'session':
        fc_syntax = "\nff =~ a*s1 + a*s2 + a*s3 + a*s4"
    elif fc_unit == 'day':
        fc_syntax = "\nff =~ day1 + day2"
        phase_encoding, day_cor = False, False

    def generate_pair_sessions(phase_encoding, day_cor, order_in_day):
        """
        Inner function for generating syntax for errors
        """
        pair_session_day = [['s1', 's2'], ['s3', 's4']] if day_cor else None
        pair_session_pe = [['s1', 's4'], ['s2', 's3']] if phase_encoding else None
        pair_session_order = [['s1', 's3'], ['s2', 's4']] if order_in_day else None
        return [pair_session_day, pair_session_pe, pair_session_order]

    error_syntax = ''
    if any([phase_encoding, day_cor, order_in_day]):
        pair_session_list = generate_pair_sessions(phase_encoding, day_cor, order_in_day)
        for pair_session in pair_session_list:
            if pair_session:
                pos00, pos01, pos10, pos11 = pair_session[0][0], pair_session[0][1], pair_session[1][0], pair_session[1][1]
                if not add_method_marker:
                    error_syntax += f'\n{pos00} ~~ {pos01}\n{pos10} ~~ {pos11}'
                else:
                    error_syntax = f'\nmf1 =~ m1 * {pos00} + m1 * {pos01}\nmf2 =~ m1 * {pos10} + m1 * {pos11}\nff ~~ 0 * mf1\nff ~~ 0 * mf2'
                    method_cor_syntax = '\nmf1 ~~ 0 * mf2' if not method_cor else ''
                    error_syntax += method_cor_syntax

    if multistate_single_trait:
        if not std_lv:
            # specify error variance of latent factors
            fc_syntax = '\no1 =~ 1*s1 + 1*s2\no2 =~ 1*s3 + 1*s4\nff =~ 1*o1 + 1*o2\nff ~~ t_var*ff\no1 ~~ o_var*o1\no2 ~~ o_var*o2\ns1 ~~ e_var1*s1\ns2 ~~ e_var2*s2\ns3 ~~ e_var3*s3\ns4 ~~ e_var4*s4'
        else:
            if diff_load:
                # not specify error variance of latent factors
                fc_syntax = '\no1 =~ a*s1 + b*s2\no2 =~ c*s3 + d*s4\nff =~ e*o1 + f*o2\ns1 ~~ e_var1*s1\ns2 ~~ e_var2*s2\ns3 ~~ e_var3*s3\ns4 ~~ e_var4*s4\na > 0\nb > 0\nc > 0\nd > 0\ne > 0\nf > 0'
            else:
            #    fc_syntax = '\no1 =~ 1*s1 + 1*s2\no2 =~ 1*s3 + 1*s4\nff =~ 1*o1 + 1*o2\ns1 ~~ e_var1*s1\ns2 ~~ e_var2*s2\ns3 ~~ e_var3*s3\ns4 ~~ e_var4*s4'
                fc_syntax = '\no1 =~ a*s1 + a*s2\no2 =~ a*s3 + a*s4\nff =~ c*o1 + c*o2\ns1 ~~ e_var1*s1\ns2 ~~ e_var2*s2\ns3 ~~ e_var3*s3\ns4 ~~ e_var4*s4'
            if phase_encoding:
                ind1, ind2, ind3, ind4 = 's1', 's4', 's2', 's3'
            elif order_in_day:
                ind1, ind2, ind3, ind4 = 's1', 's3', 's2', 's4'

            if CU:
                fc_syntax += f'\n{ind1} ~~ {ind2}\n{ind3} ~~ {ind4}'
            elif OM:
                if not m_minus_1:
                    fc_syntax += f'\nm1 =~ ml*{ind1} + ml*{ind2}\nm2 =~ ml*{ind3} + ml*{ind4}\nm1 ~~ 0*ff\nm2 ~~ 0*ff\nm1 ~~ 0*m2'
                else:
                    fc_syntax += '\nm2 =~ ml*{ind3} + ml*{ind4}\nm2 ~~ 0*ff'
    
    if bi_factor:
        fc_syntax = '\no1 =~ a*s1 + a*s2\no2 =~ a*s3 + a*s4\nff =~ c*s1 + c*s2 + c*s3 + c*s4\ns1 ~~ e_var1*s1\ns2 ~~ e_var2*s2\ns3 ~~ e_var3*s3\ns4 ~~ e_var4*s4\no1 ~~ 0*o2\nff ~~ 0*o1\nff ~~ 0*o2'

    fc_syntax += error_syntax
    if (control is not None) and (model == "model_onlyFC"):
        control_syntax = generate_model_syntax_control(control, model)
        fc_syntax += control_syntax
    return fc_syntax


def calculate_param_num(
    model: ModelFA,
    control,
    cov_cor,
    phase_encoding,
    day_cor,
    use_lavaan,
    fc_unit,
    trait_type=None,
    scale_name=None,
    remove_vars_list=None,
    return_nvars=False,
    order_in_day=False,
    add_method_marker=False,
    multistate_single_trait=False,
    bi_factor=False,
    m_minus_1=False,
    add_CU=False,
    add_OM=False,
    mean_structure=False
) -> tuple[int, int]:
    """
    Calculate a number of parameters
    Argment of cov_cor is already unnecessary. Argument cov_cor should be removed in the future.
    """
    # set number of scales or items
    num_remove = len(remove_vars_list) if remove_vars_list is not None else 0
    if trait_type == "personality":
        num_scales = 12 - num_remove
    elif trait_type == "cognition":
        num_scales = len(generate_NIH_cognition_scales(scale_name, add_comp=False)) - num_remove
    # None should be replaced
    elif trait_type == "mental":
        num_scales = len(ASR_DICT.get(scale_name))
    if not multistate_single_trait:
        if not add_method_marker:
            pe_error_num = 2 if phase_encoding else 0
            day_error_num = 2 if day_cor else 0
            order_error_num = 2 if order_in_day else 0
            add_error_num = pe_error_num + day_error_num + order_error_num
        else:
            add_error_num = 7 if any([phase_encoding, day_cor, order_in_day]) else 0
    else:
        add_error_num = 0

    if fc_unit == 'session':
        param_num_fc = 4
    elif fc_unit == 'day':
        param_num_fc = 2
    
    if model == "model_onlyFC":
        var_num = param_num_fc
        if control:
            var_num += len(control)

        # number of factor loadings (=~)
        loading_num = param_num_fc
        if multistate_single_trait:
            loading_num += 2
        if bi_factor:
            loading_num += 4
        if add_method_marker:
            loading_num += 4
        # number of variance and covariances (~~)
        vcov_num = param_num_fc + 1
        if multistate_single_trait:
            # variance of FC factors (o1, o2, and ff in param_order)
            vcov_num += 2
        if bi_factor:
            vcov_num += 5
        if add_method_marker:
            # variance of method factors (2) + variance-covariance of trait and method factors (3: ff ~~ mf1, ff ~~ mf2, mf1 ~~ mf2)
            vcov_num += 5
        if control:
            var_controls = len(control)
            comb_num = comb(len([i for i in control if 'Time' not in i]), 2)
            add_cov_time = 1 if sum(['Time' in i for i in control]) == 2 else 0
            vcov_num += comb_num + var_controls + add_cov_time
        if add_CU:
            if order_in_day:
                vcov_num += 2
        # number of intercepts (~1)
        if mean_structure:
            int_num = param_num_fc
            if multistate_single_trait:
                int_num += 3
            if control:
                int_num += len(control)
        else:
            int_num = 0

        # number of regressions (~)
        reg_num = len(control) if control else 0
        #add_time_reg_num = 2 * 2 if sum(['Time' in i for i in control]) == 2 else 0
        add_time_reg_num = 0
        reg_num += add_time_reg_num
#        param_num = (param_num_fc * 2) + 1 + add_error_num
#        var_num = param_num_fc
    
    elif model == "model_fc":
        var_num = 1 + param_num_fc
        if control:
            var_num += len(control)

        # number of factor loadings (=~)
        loading_num = 1 + param_num_fc
        if multistate_single_trait:
            loading_num += 2

        # number of variance and covariances (~~)
        vcov_num = param_num_fc + 1
        if multistate_single_trait:
            # variance of FC factors and FC-trait covariance
            vcov_num += 4 + 1
        if control:
            comb_num = comb(len(control), 2)
            var_controls = len(control)
            vcov_num += comb_num + var_controls

        # number of intercepts (~1)
        if mean_structure:
            int_num = param_num_fc + 1
            int_num += 4
            if control:
                int_num += len(control)
        else:
            int_num = 0
        # number of regressions (~)
        reg_num = len(control) * 2 if control else 0
        param_num = (param_num_fc + 1) * 2 + 3 + add_error_num
    elif model == "model_trait":
        param_num = (num_scales + 1) * 2 + 3
        var_num = num_scales + 1
    elif model in ["model_both", "model_fc_sum_trait"]:
        var_num = num_scales + param_num_fc
        if control:
            var_num += len(control)

        # number of factor loadings (=~)
        loading_num = num_scales + param_num_fc
        if multistate_single_trait:
            loading_num += 2

        # number of variance and covariances (~~)
        vcov_num = num_scales + param_num_fc
        if multistate_single_trait:
            # variance of FC factors and FC-trait covariance
            vcov_num += 4 + 1
        if control:
            comb_num = comb(len(control), 2)
            var_controls = len(control)
            vcov_num += comb_num + var_controls

        # number of intercepts (~1)
        if mean_structure:
            int_num = num_scales + param_num_fc
            int_num += 4
            if control:
                int_num += len(control)
        else:
            int_num = 0
        # number of regressions (~)
        reg_num = len(control) * 2 if control else 0
    param_num = loading_num + vcov_num + int_num + reg_num
#        param_num = (num_scales + param_num_fc) * 2 + 2 + add_error_num
#        var_num = num_scales + param_num_fc
#  
#    if multistate_single_trait:
#        param_num += 4
##        if model == 'model_both':
##            param_num += num_scales * 2
#        if add_OM:
#            if phase_encoding ^ order_in_day:
#                if not m_minus_1:
#                    param_num += 9
#                else:
#                    param_num += 4
#            elif not phase_encoding and not order_in_day:
#                pass
#            else:
#                raise Exception('PE and OrderInDay are simultaneously specified.')
#        elif add_CU:
#            if phase_encoding:
#                param_num += 2
#            if order_in_day:
#                param_num += 2
#    
#    if mean_structure:
#        param_num += num_scales + param_num_fc 
#        if multistate_single_trait:
#            param_num += 4
#
#    if control is not None:
#        comb_num = comb(len(control), 2) if cov_cor else 0
#        if model in ['model_fc', 'model_trait', 'model_both']:
#            param_num += len(control) * 2 + comb_num
#        elif model == 'model_onlyFC':
#            param_num += len(control) + comb_num
#        var_num = var_num + len(control)
#        add_num = len(control) if use_lavaan else 0
#        param_num += add_num
#        if mean_structure:
#            param_num += len(control)
    if not return_nvars:
        return param_num
    return param_num, var_num


def generate_syntax_items(scale_name, scales, remove_vars_list, equal_loadings=False):
    # a represents equality constraitns of factor loadings
    if remove_vars_list is not None:
        scales = [i for i in scales if i not in remove_vars_list]
    tf_syntax_items = ("tf =~ " + "+".join(scales)) if not len(scales) == 2 else ("tf =~ " + 'tl * ' + ' + tl *'.join(scales))
    if equal_loadings:
        tf_syntax_items = ("tf =~ " + 'tl * ' + ' + tl *'.join(scales))
    return tf_syntax_items


#def get_omega(trait_scale_name):
#    """
#    Get omega reliability to add error variance in model_fc
#    """
#    target_dir = op.join(FA_PARAMS_DIR, trait_scale_name, 'tables')
#    filenames = 


def generate_model_syntax_cog(
    model,
    scale_name,
    control,
    remove_vars_list: Union[list[str], list[None]] = [None],
    phase_encoding=True,
    day_cor=True,
    fc_unit='session',
    fix_loadings_to_one=False,
    fix_cov_to_zero=False,
    order_in_day=False,
    age_adj=False,
    multistate_single_trait=False,
    equal_loadings=False,
    add_trait_error=False
):
    """function for generating syntax for NIH cognitive toolbox"""
    if model in ['model_fc', 'model_both']:
        fc_syntax = generate_model_syntax_fc(
                model, control, phase_encoding, day_cor, fc_unit, order_in_day,
                multistate_single_trait=multistate_single_trait,
                bi_factor=bi_factor
                )
        cov_syntax = 'tf ~~ ff' if not fix_cov_to_zero else 'tf ~~ 0 * ff'
        if fc_to_trait:
            cov_syntax = 'tf ~ ff'

    if model == "model_fc":
        scales = generate_NIH_cognition_scales(
            scale_name, add_comp=True, remove_vars_list=remove_vars_list
        )
        if not add_trait_error:
            tf_syntax = "tf =~ 1*" if fix_loadings_to_one else "tf =~ "
        else:
            pass
        model_syntax = tf_syntax + scales[-1] + fc_syntax + "\n" + cov_syntax
    elif model == "model_trait":
        scales = generate_NIH_cognition_scales(
            scale_name, add_comp=False, remove_vars_list=remove_vars_list
        )
        mean_fc_syntax = "ff =~ 1 * summary_edge_fc" if fix_loadings_to_one else "ff =~ summary_edge_fc"
        model_syntax = generate_syntax_items(scale_name, scales, remove_vars_list) + "\n" + mean_fc_syntax + "\n" + cov_syntax
    elif model == "model_both":
        scales = generate_NIH_cognition_scales(
            scale_name, add_comp=False, remove_vars_list=remove_vars_list
        )
        model_syntax = generate_syntax_items(scale_name, scales, remove_vars_list, equal_loadings=equal_loadings) + fc_syntax + "\n" + cov_syntax
    elif model == "model_only_trait":
        scales = generate_NIH_cognition_scales(
            scale_name, add_comp=False, remove_vars_list=remove_vars_list, age_adj=age_adj
        )
        model_syntax = generate_syntax_items(scale_name, scales, remove_vars_list, equal_loadings=equal_loadings)

    if control is not None:
        control_syntax = generate_model_syntax_control(control, model)
        model_syntax += control_syntax

    return model_syntax


def generate_model_syntax_ffi(
    model,
    scale_name,
    ordered,
    control,
    remove_vars_list: list[Optional[str]] = [None],
    phase_encoding=True,
    use_lavaan=False,
    day_cor=True,
    fc_unit='session',
    fix_loadings_to_one=False,
    fix_cov_to_zero=False,
    order_in_day=False,
    cor_error_list:Optional[list[list]]=None,
    multistate_single_trait=False,
    equal_loadings=False,
    add_trait_error=False
):
    """generate model syntax of factor models for neo-ffi"""
    # generate items
    items = generate_ffi_subscale_item_number(scale_name, remove_vars_list=remove_vars_list)
    if model in ['model_fc', 'model_both']:
        fc_syntax = generate_model_syntax_fc(
                model, control, phase_encoding, day_cor, fc_unit, order_in_day,
                multistate_single_trait=multistate_single_trait,
                bi_factor=bi_factor,
                )
        cov_syntax = 'tf ~~ ff' if not fix_cov_to_zero else 'tf ~~ 0 * ff'
        if fc_to_trait:
            cov_syntax = 'tf ~ ff'
    # treat as ordered variables if specified
    if ordered:
        variables = " ".join(items)
        ordered_syntax = f"DEFINE(ordinal) {variables}"

    # Syntax for trait scale
    if not equal_loadings:
        trait_syntax = "tf =~ " + "+".join(items)
    else:
        trait_syntax = "tf =~ " + 'tl * ' + ' + tl *'.join(items)

    if model == "model_fc":
        if not add_trait_error:
            trait_syntax = "tf =~ 1 * Total" if fix_loadings_to_one else "tf =~ Total"
        else:
            pass
        model_syntax = trait_syntax + fc_syntax + '\n' + cov_syntax
    elif model == "model_trait":
        mean_fc_syntax = "ff =~ 1 * summary_edge_fc" if fix_loadings_to_one else 'ff =~ summary_edge_fc'
        model_syntax = "tf =~ " + "+".join(items) + "\n" + mean_fc_syntax + "\ntf ~~ ff"
        if ordered and not use_lavaan:
            model_syntax += "\n" + ordered_syntax
    elif model == "model_fc_sum_trait":
        model_syntax = "tf =~ 1*" + " + a*".join(items) + fc_syntax + "\ntf ~~ ff"
    elif model == "model_both":
        model_syntax = trait_syntax + fc_syntax + "\n" + cov_syntax
        if ordered and not use_lavaan:
            model_syntax += "\n" + ordered_syntax
    elif model == "model_only_trait":
        model_syntax = trait_syntax
        if cor_error_list is not None:
            error_syntax_all = ''
            for inner_list in cor_error_list:
                error_syntax = '\n' + inner_list[0] + ' ~~ ' + inner_list[1]
                error_syntax_all += error_syntax
            model_syntax += error_syntax_all
        if ordered:
            model_syntax += "\n" + ordered_syntax
    else:
        print(
            'Input correct model; "model_fc", "model_trait", "model_both", and.or "model_only_trait".'
        )

    if control is not None:
        control_syntax = generate_model_syntax_control(control, model)
        model_syntax += control_syntax
    return model_syntax


def generate_model_syntax_asr(
    model,
    scale_name,
    control,
    remove_vars_list: Union[list[str], list[None]] = [None],
    phase_encoding=True,
    day_cor=True,
    fc_unit='session',
    fix_loadings_to_one=False,
    fix_cov_to_zero=False,
    order_in_day=False,
    multistate_single_trait=False,
    equal_loadings=False
):
    """function for generating syntax for NIH cognitive toolbox"""
    if model in ['model_fc', 'model_both']:
        fc_syntax = generate_model_syntax_fc(
                model, control, phase_encoding, day_cor, fc_unit, order_in_day,
                multistate_single_trait=multistate_single_trait,
                bi_factor=bi_factor
                )
        cov_syntax = 'tf ~~ ff' if not fix_cov_to_zero else 'tf ~~ 0 * ff'
        if fc_to_trait:
            cov_syntax = 'tf ~ ff'
    items = ASR_DICT.get(scale_name)
    
    scales = generate_ASR_scales(scale_name, remove_vars_list=remove_vars_list)
    trait_syntax = generate_syntax_items(scale_name, scales, remove_vars_list, equal_loadings=equal_loadings)

    if model == "model_fc":
        tf_syntax = "tf =~ 1 * " if fix_loadings_to_one else 'tf =~'
        model_syntax = tf_syntax + scale_name + fc_syntax + "\n" + cov_syntax
    elif model == 'model_trait':
        mean_fc_syntax = 'ff =~ 1 * summary_edge_fc' if fix_loadings_to_one else 'ff =~ summary_edge_fc'
        model_syntax = "tf =~ " + "+".join(items) + "\n" + mean_fc_syntax + "\n" + cov_syntax
    elif model == 'model_both':
        model_syntax = trait_syntax + fc_syntax + "\n" + cov_syntax
    elif model == "model_only_trait":
        model_syntax = trait_syntax

    if control is not None:
        control_syntax = generate_model_syntax_control(control, model)
        model_syntax += control_syntax

    return model_syntax


def prepare_dataset_for_analyses(
    model,
    edge,
    control,
    cfa_standardize,
    subjects_set_for_analysis=None,
    **data_dict,
):
    """prepare pandas dataframe for model fits"""
    fc_data, trait_data, covariates_data = (
        data_dict["fc_data"],
        data_dict["trait_data"],
        data_dict["covariates_data"],
    )
    # prepare fc data at edge level
    if fc_unit == 'session':
        columns_dict = {0: "s1", 1: "s2", 2: "s3", 3: "s4"}
    elif fc_unit == 'day':
        columns_dict = {0: 'day1', 1: 'day2'}
    elif fc_unit == 'full':
        columns_dict = {0: 'summary_edge_fc'}
    fc_edge_data = (
        pd.DataFrame(fc_data[edge, :, :])
        .rename(columns=columns_dict)
        .reset_index()
        .drop("index", axis=1)
    )
    # concatenate fc and trait data
    data_for_cfa = pd.concat([fc_edge_data, trait_data], axis=1)
    # calculate mean of fc when necessary
    if model in ["model_trait", "mean"]:
        if fc_unit in ['day', 'session']:
            data_for_cfa["summary_edge_fc"] = (
                data_for_cfa["s1"]
                + data_for_cfa["s2"]
                + data_for_cfa["s3"]
                + data_for_cfa["s4"]
            ) / 4

    # merge data of covariates
    if control is not None:
        data_for_cfa = pd.merge(data_for_cfa, covariates_data)
    if cfa_standardize:
        scaler = StandardScaler()
        std_columns = data_for_cfa.select_dtypes(include=["float64", "int64"]).columns
        data_for_cfa[std_columns] = scaler.fit_transform(data_for_cfa[std_columns])
    return data_for_cfa


def apply_model_fit(
        model_syntax,
        data_for_cfa,
        model_fit_obj,
        model=None, 
        use_lavaan=False,
        multistate_single_trait=False,
        std_lv=True,
        mean_structure=False
        ):
    """
    function for model fits using semopy or lavaan package
    """
    # fit one model
    if not use_lavaan:
        model_cfa = Model(model_syntax)
        model_cfa.fit(data_for_cfa, obj=model_fit_obj)  # take long time
        fit_model = model_cfa
    else:
        # generate dataframe for analyses using R
        with (ro.default_converter + pandas2ri.converter).context():
            data_for_cfa_r = ro.conversion.get_conversion().py2rpy(data_for_cfa)
        globalenv["data_for_cfa_r"] = data_for_cfa_r
        # generate code for categorical CFA
        if (
            (model is not None)
            and (model in ["model_trait", "model_both"])
            and ordered
            and (trait_type == "personality")
        ):
            ordered_vars = (
                "c("
                + ", ".join(
                    [f'"{i}"' for i in generate_ffi_subscale_item_number(scale_name)]
                )
                + ")"
            )
            ordered_code = f"data_for_cfa_r[, {ordered_vars}] <- lapply(data_for_cfa_r[, {ordered_vars}], ordered)"
            ro.r(ordered_code)
            ordered_code_in_cfa = f", ordered = {ordered_vars}"
            model_fit_obj = "WLSMV"
        else:
            ordered_code_in_cfa = ""

        try: 
            iter_max
        except NameError:
            convergence_syntax = '' 
        else:
            convergence_syntax = f', control = list(iter.max = {iter_max}, rel.tol = {rel_tol})'
        std_syntax = ', std.lv = TRUE' if std_lv else ''
        mean_structure_syntax = ', meanstructure = TRUE' if mean_structure else ''
        cfa_code = f'cfa("{model_syntax}", data = data_for_cfa_r, estimator = "{model_fit_obj}"{std_syntax}{ordered_code_in_cfa}{mean_structure_syntax}' + convergence_syntax + ')'
        fit_model = ro.r(cfa_code)
    return fit_model


def calculate_srmr(fit_model, input_df, out_diff_mat=False):
    """
    function for calculating SRMR
    referring to https://note.com/k_fukunaka/n/n994613904942
    """
    model_mat = Model.calc_sigma(fit_model)
    sigma_hat = model_mat[0]
    rho_hat = covariance_to_correlation(sigma_hat)
    input_df = input_df[fit_model.vars["observed"]]
    r = input_df.corr().to_numpy()
    # calculate SRMR
    p = r.shape[0]
    diff_mat = r - rho_hat
    A = (r - rho_hat) ** 2
    num = (A.sum() - np.diag(A).sum()) / 2
    SRMR = np.power(2 * num / (p * (p + 1)), 0.5)
    if out_diff_mat:
        return SRMR, diff_mat
    return SRMR


def get_model_fits(
    fit_model, dtype_memmap, use_lavaan=False
) -> NDArray[Shape["15"], Float]:
    """get model fits"""
    if not use_lavaan:
        model_fits = calc_stats(fit_model).T["Value"].to_numpy(dtype=dtype_memmap)
    else:
        globalenv["fit_model"] = fit_model
        fit_code = f"fitMeasures(fit_model, c('df', 'baseline.df', 'chisq', 'pvalue', 'baseline.chisq', 'cfi', 'gfi', 'agfi', 'nfi', 'tli', 'rmsea', 'aic', 'bic', 'logl', 'srmr'))"
        model_fits = ro.r(fit_code)
        model_fits = np.asarray(model_fits, dtype=dtype_memmap)
    return model_fits


def get_fscore(
    fit_model, data, use_lavaan, dtype_memmap
) -> NDArray[Shape["Sample_n, Num_of_columns_of_data"], Float]:
    """
    get factor scores
    """
    if not use_lavaan:
        f_score = fit_model.predict_factors(data)
    else:
        fscore_code = "lavPredict(fit_model)"
        globalenv["fit_model"] = fit_model
        f_score = ro.r(fscore_code)
    return np.array(f_score, dtype=dtype_memmap)


def get_params_sub(
        fit_model,
        se_robust: bool,
        use_lavaan: bool,
        get_std_solutions=True,
        ordered=False
        ):
    """get parameters from semopy outputs"""
    if not use_lavaan:
        param_summary = fit_model.inspect(std_est=True, se_robust=se_robust)
    else:
        param_code = "standardizedSolution(fit_model)" if get_std_solutions else "parameterEstimates(fit_model)" 
        globalenv["fit_model"] = fit_model
        param_summary = ro.r(param_code)
        with (ro.default_converter + pandas2ri.converter).context():
            param_summary = ro.conversion.get_conversion().rpy2py(param_summary)
        if ordered:
            param_summary.query("op.isin(['=~', '~', '~~', '~1'])", inplace=True)
    print(param_summary)
    return param_summary


def get_params_from_model(
        fit_model,
        dtype_memmap, 
        se_robust: bool, 
        use_lavaan: bool,
        return_array=True,
        save_variables=False,
        get_std_solutions=True,
        ordered=False
        ) -> Union[NDArray, pd.DataFrame]:
    """
    get parameters from models
    get_std_solutions controls column selection of lavaan outputs
    get_std_solutions selects parameterestimates() or standardizedsolution() in lavaan
    """
    param_summary = get_params_sub(fit_model, se_robust, use_lavaan, get_std_solutions=get_std_solutions, ordered=ordered)
    if not use_lavaan:
        param_summary = param_summary[
            ["lval", "op", "rval", "Est. Std", "Estimate", "Std. Err"]
        ].sort_values(["op", "lval", "rval"])
        if not save_variables:
            param_summary = param_summary[["Est. Std", "Estimate", "Std. Err"]]
        param_summary[param_summary == "-"] = np.nan
    else:
        if not save_variables:
            est_list = ["se", "pvalue", "ci.lower", "ci.upper"]
            est_column = ['est'] if not get_std_solutions else ['est.std']
            param_summary = param_summary[est_column + est_list]
    if return_array:
        param_summary = param_summary.to_numpy(dtype=dtype_memmap)
    return param_summary


def generate_param_order_filename(
    control: list[str],
    model: str,
    cov_cor: bool,
    phase_encoding: bool,
    day_cor: bool,
    use_lavaan: bool,
    fc_unit: str,
    trait_type: str = None,
    scale_name: str = None,
    ordered_ffi=False,
    drop_vars_list=None,
    fix_loadings_to_one=False,
    fix_cov_to_zero=False,
    order_in_day=False,
    add_method_marker=False,
    multistate_single_trait=False,
    single_trait=False,
    bi_factor=False,
    m_minus_1=False,
    CU=False,
    OM=False,
    mean_structure=False
) -> str:
    """function for generating filename of parameter order"""
    if trait_type is not None:
        folder = select_folder_from_trait(trait_type)
    
#    if trait_type == 'personality' and ordered_ffi:
#        polycor_suffix = '_PolyCor'
#    else:
#        polycor_suffix = ''

    control_str = "_".join(control) if control is not None else None
    
    cov_cor_suffix = "_CovCor" if cov_cor else ""
    
    if not multistate_single_trait:
        pe_suffix = '_PEErrorCor' if phase_encoding else ''
    else:
        pe_suffix = '_PEFactor' if phase_encoding else ''
    
    lavaan_suffix = "_lavaan" if use_lavaan else ""
    daycor_suffix = '_DayErrorCor' if day_cor else ''
    
    if CU:
        order_suffix = '_OrderErrorCor' if order_in_day else ''
    else:
        order_suffix = '_OrderFactor' if order_in_day else ''
    
    m1_suffix = '_mMinus1' if m_minus_1 else ''
    cu_suffix = '_CU' if CU else ''
    om_suffix = '_OM' if OM else ''
    fixed_suffix = '_FixedLoad' if fix_loadings_to_one else ''
    fix_cov_zero_suffix = '_FixCovZero' if fix_cov_to_zero else ''
    method_marker_suffix = '_addMarker' if add_method_marker else ''
    st_suffix = '_ST' if single_trait else ''
    msst_suffix = '_MSST' if multistate_single_trait else ''
    bifactor_suffix = '_Bifactor' if bi_factor else ''
    mean_str_suffix = '_MeanStr' if mean_structure else ''
    fc_to_trait_suffix = '_FCToTrait' if fc_to_trait else ''
    if drop_vars_list is not None:
        if trait_type == 'personality':
            drop_vars_list = [i.replace('NEORAW_', '') for i in drop_vars_list]
        drop_suffix = '_drop_' + '_'.join(drop_vars_list)
    else:
        drop_suffix = ''
    
    filename = f"{model}_controlling_{control_str}{cov_cor_suffix}{lavaan_suffix}{cu_suffix}{om_suffix}{pe_suffix}{daycor_suffix}{m1_suffix}{method_marker_suffix}{st_suffix}{msst_suffix}{bifactor_suffix}{mean_str_suffix}{fc_to_trait_suffix}{order_suffix}_{fc_unit}"
    if model in ['model_trait', 'model_both']:
        # polycor_suffix may be added
        filename += f"{drop_suffix}{fixed_suffix}{fix_cov_zero_suffix}.csv"
    elif model == 'model_fc':
        filename += '.csv'
#    elif model == 'model_fc':
#        filename = f"{model}_controlling_{control_str}{cov_cor_suffix}{lavaan_suffix}{pe_suffix}{daycor_suffix}{order_suffix}_{fc_unit}{fixed_suffix}{fix_cov_zero_suffix}.csv"
    elif model == 'model_onlyFC':
        filename += '.csv' 
    
    if trait_type is not None:
        param_order_file = op.join(FA_PARAMS_DIR, folder, f"{scale_name}_{filename}")
    else:
        param_order_file = op.join(FA_PARAMS_DIR, "reliability", filename)
    
    return param_order_file


class ParamPositionDict(TypedDict):
    cor_position: int
    fc_load_positions: list[int]
    trait_load_positions: list[int]
    fc_error_positions: list[int]
    trait_error_positions: list[int]


def get_param_positions(
    param_order_file, use_lavaan: bool, control, model, fc_unit
) -> ParamPositionDict:
    """
    Function for returning index of estimate of inter-factor correlation
    """
    param_order = pd.read_csv(param_order_file).reset_index()
    
    if fc_unit == 'session':
        fc_index = ["s" + str(i) for i in range(1, 5)] + ["summary_edge_fc"]
    elif fc_unit == 'day':
        fc_index = ['day1', 'day2']
    
    fc_index_sessions = ["s1", "s2", "s3", "s4"]
    fc_index_day1 = ["s1", "s2"]
    fc_index_day2 = ["s3", "s4"]

    method_factors = ['mf1', 'mf2']

    fc_and_factor_index = fc_index + ["tf", "ff"]
    fc_factor_control_index = fc_and_factor_index + control if control else fc_and_factor_index
    trait_exist = model in ["model_fc", "model_trait", "model_both"]
    
    if not use_lavaan:
        if trait_exist:
            param_position_dict: ParamPositionDict = {
                # get index of inter-factor correlation and save order of parameters once in iteration
                "cor_position": param_order.query(
                    'lval == "tf" & rval == "ff" & op == "~~"'
                ).index.item(),
                # get index of loading(s) of FC indicator(s)
                "fc_load_positions": param_order.query(
                    'op == "~" & rval == "ff"'
                ).index.values.tolist(),
                # get index of loading(s) of trait indicator(s)
                "trait_load_positions": param_order.query(
                    'op == "~" & rval == "tf"'
                ).index.values.tolist(),
                # get index of error(s) of FC indicator(s)
                "fc_error_positions": param_order.query(
                    'lval == rval & op == "~~" & lval in @fc_index'
                ).index.values.tolist(),
                # get index of errors of trait indicator(s)
                "trait_error_positions": param_order.query(
                    'lval == rval & op == "~~" & lval not in @fc_and_factor_index'
                ).index.values.tolist(),
            }
        elif "day" in model:
            param_position_dict = {
                # get index of inter-factor correlation and save order of parameters once in iteration
                "cor_position": param_order.query(
                    'lval == "fc_d1" & rval == "fc_d2" & op == "~~"'
                ).index.item(),
                # get index of loading(s) of FC indicators in day 1
                "fc_load_positions_day1": param_order.query(
                    'op == "~" & rval == "fc_d1"'
                ).index.values.tolist(),
                # get index of loading(s) of FC indicators in day 2
                "fc_load_positions_day2": param_order.query(
                    'op == "~" & rval == "fc_d2"'
                ).index.values.tolist(),
                # get index of error(s) of FC indicators in day 1
                "fc_error_positions_day1": param_order.query(
                    'lval == rval & op == "~~" & lval in @fc_index_day1'
                ).index.values.tolist(),
                # get index of errors of FC indicator in day 2
                "fc_error_positions_day2": param_order.query(
                    'lval == rval & op == "~~" & lval in @fc_index_day2'
                ).index.values.tolist(),
            }
        elif model == "model_onlyFC":
            param_position_dict = {
                # get index of loading(s) of FC indicators
                "fc_load_positions": param_order.query(
                    'op == "~" & rval == "fc"'
                ).index.values.tolist(),
                # get index of error(s) of FC indicators
                "fc_error_positions": param_order.query(
                    'lval == rval & op == "~~" & lval in @fc_index_sessions'
                ).index.values.tolist(),
            }

    else:
        if trait_exist:
            cor_position_op = '~~' if not fc_to_trait else '~'
            param_position_dict: ParamPositionDict = {
                # get index of inter-factor correlation and save order of parameters once in iteration
                "cor_position": param_order.query(
                    f'lhs == "tf" & rhs == "ff" & op == "{cor_position_op}"'
                ).index.item(),
                # get index of loading(s) of FC indicator(s)
                "fc_load_positions": param_order.query(
                    'op == "=~" & lhs == "ff"'
                ).index.values.tolist(),
                # get index of loading(s) of trait indicator(s)
                "trait_load_positions": param_order.query(
                    'op == "=~" & lhs == "tf"'
                ).index.values.tolist(),
                # get index of error(s) of FC indicator(s)
                "fc_error_positions": param_order.query(
                    'lhs == rhs & op == "~~" & lhs in @fc_index'
                ).index.values.tolist(),
                # get index of errors of trait indicator(s)
                "trait_error_positions": param_order.query(
                    'lhs == rhs & op == "~~" & lhs not in @fc_factor_control_index'
                ).index.values.tolist(),
            }
        elif "session" in model or 'onlyFC' in model:
            param_position_dict = {
                # get index of loading(s) of FC indicators
                "fc_load_positions": param_order.query(
                    'lhs == "ff" & rhs in @fc_index_sessions'
                ).index.values.tolist(),
                # get index of error(s) of FC indicators
                "fc_error_positions": param_order.query(
                    'lhs == rhs & op == "~~" & lhs in @fc_index_sessions'
                ).index.values.tolist(),
                # get index of covariances of FC indicators
                "fc_cov_positions": param_order.query(
                    'lhs != rhs & lhs in @fc_index_sessions & rhs in @fc_index_sessions'
                ).index.values.tolist(),
                # get_index of loadings of method factors
                "fc_marker_positions": param_order.query(
                    'lhs in @method_factors & rhs in @fc_index_sessions'
                    ).index.tolist()
            }
    return param_position_dict


def try_model_fit_inspect(
        model,
        model_syntax,
        edge_input, 
        data_dict,
        **kwargs
        ):
    """
    Function for inspecting fitted model
    """
    data_for_fa_test = prepare_dataset_for_analyses(
        model,
        edge_input,
        control,
        cfa_standardize,
        **data_dict,
    )
    # model fit
    fit_model = apply_model_fit(
            model_syntax, 
            data_for_fa_test,
            model_fit_obj,
            model,
            use_lavaan,
            multistate_single_trait=multistate_single_trait,
            mean_structure=mean_structure
            )
    # get parameter order
    try:
        # sort order should match param_summary_sub()
        if not use_lavaan:
            param_order = (
                fit_model.inspect()[["lval", "op", "rval"]]
                .sort_values(by=["op", "lval", "rval"])
                .reset_index(drop=True)
            )
        else:
            param_order = get_params_sub(
                fit_model, kwargs["se_robust"], kwargs["use_lavaan"], kwargs['multistate_single_trait'], ordered=ordered
            )[["lhs", "op", "rhs"]]
    except (np.linalg.LinAlgError, rpy2.rinterface_lib.embedded.RRuntimeError):
        param_order = None
    return param_order


def save_param_order(model, model_syntax, data_dict, over_write=True, **kwargs):
    """function for generating and saving order of model parameters"""
    param_order_file = generate_param_order_filename(
        control,
        model,
        kwargs["cov_cor"],
        kwargs["phase_encoding"],
        kwargs['day_cor'],
        kwargs["use_lavaan"],
        kwargs['fc_unit'],
        trait_type=trait_type,
        scale_name=scale_name,
        drop_vars_list=kwargs['remove_vars_list'],
        fix_loadings_to_one=kwargs['fix_loadings_to_one'],
        fix_cov_to_zero=kwargs.get('fix_cov_to_zero'),
        order_in_day=order_in_day,
        ordered_ffi=ordered,
        add_method_marker=add_method_marker,
        single_trait=single_trait,
        multistate_single_trait=multistate_single_trait,
        bi_factor=bi_factor,
        m_minus_1=m_minus_1,
        OM=OM,
        CU=CU,
        mean_structure=mean_structure
    )
    # try to generate parameter order if parameter_order_file does not exist
    if not over_write:
        if not op.isfile(param_order_file):
            param_order = None
            for edge in range(10):
                param_order = try_model_fit_inspect(
                    model, model_syntax, edge, data_dict, **kwargs
                )
                if param_order is not None:
                    break
            # save parameter order file
            param_order.to_csv(param_order_file, index=False)
    else:
        param_order = None
        if edge_start == 0:
            for edge in range(50):
                param_order = try_model_fit_inspect(
                    model, model_syntax, edge, data_dict, **kwargs
                )
                if param_order is not None:
                    break
            # save parameter order file
            param_order.to_csv(param_order_file)


def get_residual_cors(
        fit_model, 
        return_array=True,
        ) -> Union[NDArray, pd.DataFrame]:
    """
    function for getting residual correlations used in analyses using lavaan
    """
    if return_array:
        res_code = "lavResiduals(fit_model, type = 'cor')['cov']"
    else:
        res_code = "data.frame(lavResiduals(fit_model, type = 'cor')['cov'])"
    globalenv["fit_model"] = fit_model
    res_cor_mat = ro.r(res_code)
    #with (ro.default_converter + pandas2ri.converter).context():
    #    res_cor_mat = ro.conversion.get_conversion().rpy2py(res_cor_mat)
    if return_array:
        res_cor_mat = np.squeeze(np.asarray(res_cor_mat))
    return res_cor_mat


def get_model_implied_vcov_mat(
        fit_model, 
        return_array=True,
        ) -> Union[NDArray, pd.DataFrame]:
    """
    function for getting residual correlations used in analyses using lavaan
    """
    code = "fitted(fit_model)$cov"
    if return_array:
        code_input = code
    else:
        code_input = f"data.frame({code})"
    globalenv["fit_model"] = fit_model
    vcov_mat = ro.r(code_input)
    #with (ro.default_converter + pandas2ri.converter).context():
    #    res_cor_mat = ro.conversion.get_conversion().rpy2py(res_cor_mat)
    if return_array:
        vcov_mat = np.squeeze(np.asarray(vcov_mat))
    return vcov_mat


def get_MIs(fit_model):
    """get modification indices"""
    mi_code = 'modificationindices(fit_model)'
    globalenv['fit_model'] = fit_model
    mi_mat = ro.r(mi_code)
    return mi_mat


def try_and_get_model_outputs(
    fit_model, 
    param_num,
    edge,
    model,
    input_df: pd.DataFrame,
    var_num,
    **kwargs
) -> tuple[NDArray, NDArray, pd.DataFrame]:
    """function for trying to get model fits and parameters"""
    try:
        # get model fits
        model_fits = get_model_fits(
            fit_model, kwargs["dtype_memmap"], kwargs["use_lavaan"]
        )
        # get summary of parameters
        param_summary = get_params_from_model(
            fit_model, 
            kwargs["dtype_memmap"], 
            kwargs["se_robust"], 
            kwargs["use_lavaan"], 
            get_std_solutions=kwargs.get('get_std_solutions'),
            ordered=ordered
        )
        # get factor scores
        fscores = get_fscore(
            fit_model, input_df, kwargs["use_lavaan"], kwargs["dtype_memmap"]
        )
        if not kwargs["use_lavaan"]:
            srmr, cor_res_mat = calculate_srmr(fit_model, input_df, out_diff_mat=True)
        else:
            cor_res_mat = get_residual_cors(fit_model)
            model_vcov_mat = get_model_implied_vcov_mat(fit_model)
        # add SRMR
        if not kwargs["use_lavaan"]:
            model_fits = np.append(model_fits, srmr)
    except (np.linalg.LinAlgError, rpy2.rinterface_lib.embedded.RRuntimeError) as err:
        if "day" in model or "session" in model:
            print(edge, model, err)
        else:
            print(scale_name, edge, model, err)
        # parameters
        ncol = 3 if not kwargs["use_lavaan"] else 5
        param_summary = np.empty(shape=(param_num, ncol))
        param_summary[:] = np.nan
        # model fits
        model_fits = np.array([np.nan] * 15)
        # factor score estimates
        if trait_type:
            fscores_ndim2 = 2
        else:
            fscores_ndim2 = 1
        if multistate_single_trait:
            fscores_ndim2 += 2
        if add_method_marker:
            fscores_ndim2 += 2
#            if phase_encoding:
#                fscores_ndim2 += 2
#            if order_in_day:
#                fscores_ndim2 += 2
        fscores = np.empty(shape=(input_df.shape[0], fscores_ndim2))
        fscores[:] = np.nan

        def create_empty_mat():
            empty_mat = np.empty(shape=(var_num, var_num))
            empty_mat[:] = np.nan
            return empty_mat

        # residuals
        cor_res_mat = create_empty_mat()
        # model implied variance covariance matrix
        model_vcov_mat = create_empty_mat()
    return param_summary, model_fits, fscores, cor_res_mat, model_vcov_mat


def get_std_parameter_estimates(param_summary, param_position_dict, model):
    """
    function for get zero-order correlation coefficient
    (standardized parameter estimates of inter-factor correlation)
    """
    if model in MODEL_TRAIT:
        param_estimates_dict = {
            "cor_estimate": param_summary[int(param_position_dict["cor_position"]), 0],
            "fc_loadings": param_summary[
                [int(i) for i in param_position_dict["fc_load_positions"]], 0
            ],
            "trait_loadings": param_summary[
                [int(i) for i in param_position_dict["trait_load_positions"]], 0
            ],
            "fc_error_vars": param_summary[
                [int(i) for i in param_position_dict["fc_error_positions"]], 0
            ],
            "trait_error_vars": param_summary[
                [int(i) for i in param_position_dict["trait_error_positions"]], 0
            ],
            "fc_error_ses": param_summary[
                [int(i) for i in param_position_dict["fc_error_positions"]], 2
            ],
            "trait_error_ses": param_summary[
                [int(i) for i in param_position_dict["trait_error_positions"]], 2
            ],
        }
    if model == "model_onlyFC":
        param_estimates_dict = {
            "fc_loadings": param_summary[
                [int(i) for i in param_position_dict["fc_load_positions"]], 0
            ],
            "fc_error_vars": param_summary[
                [int(i) for i in param_position_dict["fc_error_positions"]], 0
            ],
            "fc_error_ses": param_summary[
                [int(i) for i in param_position_dict["fc_error_positions"]], 2
            ],
        }
    return param_estimates_dict


def edge_loop_for_models(
    edge,
    model,
    model_index,
    model_syntax,
    param_num,
    var_num,
    param_position_dict,
    data_dict,
    logger_dict,
    memmap_dict,
    string_dict,
    **kwargs,
):
    """loop of edges for factor models"""
    # prepare dataset for analyses
    data_for_cfa = prepare_dataset_for_analyses(
        model,
        edge,
        control,
        cfa_standardize,
        **data_dict,
    )
    if save_cfa_data:
        save_dir = op.join(ATLAS_DIR, trait_scale_name, scale_name, "data")
        os.makedirs(save_dir, exist_ok=True)
        data_for_cfa.to_csv(
            op.join(save_dir, f"Edge_{edge}_{trait_scale_name}_{scale_name}.csv"),
            index=False,
        )
    # calculate sample size of analysis
    sample_n = data_for_cfa.shape[0]
    print(
        f"Fold: {kwargs['kfold_i']}, {model}, Edge: {edge}, N: {sample_n}, {trait_type}: {scale_name}"
    )
    # fit model
    fit_model = apply_model_fit(
        model_syntax,
        data_for_cfa,
        model_fit_obj,
        model,
        use_lavaan,
        multistate_single_trait,
        std_lv=std_lv,
        mean_structure=mean_structure
    )

    # set logger
    if not use_lavaan:

        class ContextFilter(Filter):
            """
            This is a filter which injects contextual information into the log.
            """

            def filter(self, record):
                record.gsr_type = string_dict["gsr_type"]
                record.scale_name = scale_name
                record.edge = edge
                record.Model = model
                return True

        logger, file_handler = logger_dict["logger"], logger_dict["file_handler"]
        logger.addFilter(ContextFilter())
        format_log = Formatter(
            "%(gsr_type)s,%(scale_name)s,%(edge)s,%(Model)s,%(levelname)s,%(message)s"
        )
        file_handler.setFormatter(format_log)
        logger.addHandler(file_handler)

    # get unstandardized and standardized parameter estimates and unstandardized standard errors and model fits
    param_summary, model_fits, fscores, std_residuals, model_vcov_mat = try_and_get_model_outputs(
        fit_model, param_num, edge, model, data_for_cfa, var_num, **kwargs
    )
    # get standardized parameter estimates including inter-factor correlation and factor loadings of FC and trait
    param_estimates_dict = get_std_parameter_estimates(
        param_summary, param_position_dict, model
    )

    edge_input = calc_edge_input(edge, edge_total, edge_divide_n)
    fa_model_index = model_index - len([i for i in model_type if not "model" in i])
    if not model == "model_onlyFC":
        memmap_dict["cor_memmap"][edge_input, model_index] = param_estimates_dict[
            "cor_estimate"
        ]
    if edge == edge_start:
        print(f'shape of param_summary is {param_summary.shape}.')
        print(f'shape of param_memmap is {memmap_dict["param_memmap"].shape}.')
    memmap_dict["param_memmap"][edge_input, :, :] = param_summary
    memmap_dict["fit_memmap"][edge_input, :, fa_model_index] = model_fits
    # may cause error
    memmap_dict["fscore_memmap"][edge_input, :, :] = fscores
    memmap_dict["resid_memmap"][edge_input, ...] = std_residuals
    memmap_dict["model_vcov_memmap"][edge_input, ...] = model_vcov_mat


def calc_partial_cors(df, cor_type, control):
    """calculate partial correlations."""
    control_int = [i for i in control if i in ['age', 'gender']]
    df[control_int] = df[control_int].astype(int)
    df['scale_score'] = df['scale_score'].astype(float)
    if cor_type == "pearson":
        cor = partial_corr(
            data=df, x="fc_score", y="scale_score", covar=control
        ).r.pearson.item()
    elif cor_type == "spearman":
        cor = partial_corr(
            data=df,
            x="fc_score",
            y="scale_score",
            covar=control,
            method="spearman",
        ).r.spearman.item()
    else:
        cor = np.nan
        print('Input cor_type ("pearson" or "spearman").')
    return cor


def calc_cor(data, cor_type):
    """calculate zero-order correlations"""
    if cor_type == "pearson":
        cor = np.corrcoef(data["fc_score"], data["scale_score"])[0, 1]
    elif cor_type == "spearman":
        cor = spearmanr(data["fc_score"], data["scale_score"])[0]
    return cor


def calc_edge_input(edge, edge_total, edge_divide_n):
    """
    Generate position of edge input
    """
    return int(edge % (edge_total / edge_divide_n))


def edge_loop_for_nomodels(
    kfold_i, model, model_index, edge, cor_memmap, data_dict, **kwargs
) -> None:
    """loop of edges where correlation is estimated without using models"""

    data_for_cfa = prepare_dataset_for_analyses(
        model,
        edge,
        control,
        cfa_standardize,
        **data_dict,
    )
    sample_n = data_for_cfa.shape[0]
    print(
        f"Fold: {kfold_i}, {model}, Edge: {edge}, N: {sample_n}, {trait_type}: {scale_name}"
    )

    if model == "mean":
        data_for_cfa.rename(columns={"summary_edge_fc": "fc_score"}, inplace=True)
    elif model == "pca":
        fc_data = data_dict["fc_data"]
        data_for_cfa["fc_score"], eigen_vecs = pca(
            fc_data[edge, :, :], pca_z_scaling=pca_z_scaling
        )
        if all(eigen_vecs < 0):
            data_for_cfa["fc_score"] = -data_for_cfa["fc_score"]

    cor = (
        calc_partial_cors(data_for_cfa, cor_type, control)
        if control is not None
        else calc_cor(data_for_cfa, cor_type)
    )
    print(cor)
    cor_memmap[calc_edge_input(edge, edge_total, edge_divide_n), model_index] = cor


def set_mean_pc_scores_of_trait(
    trait_type, trait_data, model, scale_name, pca_z_scaling
):
    """function for setting mean and/or calculating pc scores of trait data"""

    # set scale name of total score
    if trait_type == "personality":
        total_score_name = "Total"
    elif trait_type == "cognition":
        scale_names = generate_NIH_cognition_scales(scale_name, add_comp=True)
        total_score_name = scale_names[-1]
    elif trait_type == "mental":
        total_score_name = scale_name
    if model == "mean":
        trait_data["scale_score"] = trait_data[total_score_name].copy()
    elif model == "pca":
        # conduct pca and calculate pca scores and eigenvectors
        trait_data["scale_score"], eigen_vecs = pca(
            trait_data.drop(["Subject", total_score_name], axis=1), pca_z_scaling
        )
        # adjust sign of scores based on eigenvectors
        if all(eigen_vecs < 0):
            trait_data["scale_score"] = -trait_data["scale_score"]

    return trait_data


def generate_model_syntax_totally(model, **kwargs):
    """
    function for generating model syntax from trait type
    """
    if trait_type == "personality":
        model_syntax = generate_model_syntax_ffi(
            model,
            scale_name,
            ordered, control, 
            use_lavaan=use_lavaan, 
            phase_encoding=phase_encoding, 
            day_cor=day_cor, 
            fc_unit=fc_unit, 
            remove_vars_list=remove_vars_list,
            fix_loadings_to_one=fix_loadings_to_one,
            fix_cov_to_zero=fix_cov_to_zero,
            multistate_single_trait=multistate_single_trait,
            equal_loadings=equal_loadings,
        )
    elif trait_type == "cognition":
        model_syntax = generate_model_syntax_cog(
                model, 
                scale_name,
                control, 
                remove_vars_list,
                phase_encoding,
                day_cor, 
                fc_unit,
                fix_loadings_to_one,
                fix_cov_to_zero,
                multistate_single_trait=multistate_single_trait,
                equal_loadings=equal_loadings,
                )
    elif trait_type == "mental":
        model_syntax = generate_model_syntax_asr(
                model, 
                scale_name, 
                control, 
                remove_vars_list, 
                phase_encoding, 
                day_cor, 
                fc_unit, 
                fix_loadings_to_one, 
                fix_cov_to_zero,
                multistate_single_trait=multistate_single_trait,
                equal_loadings=equal_loadings,
                )
    elif trait_type is None:
        model_syntax = generate_model_syntax_fc(
                model, control, phase_encoding, day_cor, fc_unit, order_in_day, 
                multistate_single_trait=multistate_single_trait, 
                bi_factor=bi_factor,
                add_method_marker=add_method_marker
                )
    print(model_syntax)
    return model_syntax


def get_scale_name_from_trait(
        trait_type: TraitType,
        publication=False,
        drop=False,
        drop_subscale=False
        ) -> TraitScaleName:
    """function for selecting a folder saving outputs"""
    if trait_type == "personality":
        heading = 'c'
        trait_scale_name = "NEO_FFI" if not publication else 'NEO-FFI'
        if publication and drop:
            trait_scale_name = '(II) NEO-FFI'
        if publication and drop_subscale:
            trait_scale_name = '(II) Openness scale in the NEO-FFI'
    elif trait_type == "cognition":
        heading = 'a'
        trait_scale_name = "NIH_Cognition" if not publication else 'NIH Toolbox'
        if publication and drop_subscale:
            trait_scale_name = '(I) Fluid scale in the NIH Toolbox'
    elif trait_type == "mental":
        heading = 'b'
        trait_scale_name = "ASR" if not publication else 'ASR'
    # modified in 2024.11.29
    elif trait_type is None:
        trait_scale_name = 'reliability'
    else:
        raise Exception(f'The follwing input was passed: {trait_type}')
    return trait_scale_name


def preprocess_for_inner_edge_loop_for_models(
    model, data_dict, string_dict, memmap_dict, num_iter, sample_n, **kwargs
):
    """
    function for conducting preprocessing before looping over fa models
    """
    # set number of parameters
    model_syntax = generate_model_syntax_totally(model, **input_params)
    param_num, var_num = calculate_param_num(
        model,
        control,
        cov_cor,
        phase_encoding,
        day_cor,
        use_lavaan,
        fc_unit,
        remove_vars_list=remove_vars_list,
        trait_type=trait_type,
        scale_name=scale_name,
        return_nvars=True,
        order_in_day=order_in_day,
        add_method_marker=add_method_marker,
        multistate_single_trait=multistate_single_trait,
        bi_factor=bi_factor,
        m_minus_1=m_minus_1,
        add_CU=CU,
        add_OM=OM,
        mean_structure=mean_structure
    )
    print(f'Generating memmap files with {param_num} parameters.')
    param_memmap = generate_param_memmap(
        model,
        param_num,
        num_iter,
        string_dict["suffix"],
        **input_params,
    )
    resid_memmap = generate_resid_memmap(
        model,
        var_num,
        num_iter,
        string_dict["suffix"],
        **input_params,
    )
    model_implied_vcov_memmap = generate_model_implied_vcov_memmap(
        model,
        var_num,
        num_iter,
        string_dict["suffix"],
        **input_params,
    )
    fscore_memmap = generate_fscore_memmap(
        model, 
        sample_n, 
        num_iter, 
        string_dict["suffix"],
        add_method_marker=add_method_marker,
        **input_params
    )
    memmap_dict["param_memmap"] = param_memmap
    memmap_dict["resid_memmap"] = resid_memmap
    memmap_dict['model_vcov_memmap'] = model_implied_vcov_memmap
    memmap_dict["fscore_memmap"] = fscore_memmap
    # save parameter order file if it does not exist
    save_param_order(model, model_syntax, data_dict, **input_params)
    # get a dictionary of parameter positions
    param_order_file = generate_param_order_filename(
        control, 
        model, 
        cov_cor,
        phase_encoding,
        day_cor,
        use_lavaan,
        fc_unit,
        trait_type,
        scale_name,
        ordered,
        remove_vars_list,
        fix_loadings_to_one,
        order_in_day,
        single_trait=single_trait,
        multistate_single_trait=multistate_single_trait,
        bi_factor=bi_factor,
        mean_structure=mean_structure
    )
    param_position_dict = get_param_positions(
        param_order_file, use_lavaan, control, model, fc_unit
    )
    return model_syntax, param_num, var_num, memmap_dict, param_position_dict


def model_loop(**kwargs):
    """function for conducting model looping"""
    # model loops
    for model_index, model in enumerate(model_type):
        print(f'Start processing {model}')
        if "model_" in model:
            (
                model_syntax,
                param_num,
                var_num,
                memmap_dict,
                param_position_dict,
            ) = preprocess_for_inner_edge_loop_for_models(
                model,
                kwargs["data_dict"],
                kwargs["string_dict"],
                kwargs["memmap_dict"],
                kwargs["num_iter"],
                kwargs["sample_n"],
                **input_params,
            )
            other_dict = {
                "kfold_i": kwargs["kfold_i"],
                "model": model,
                "model_syntax": model_syntax,
                "model_index": model_index,
                "param_position_dict": param_position_dict,
                "param_num": param_num,
                "var_num": var_num,
            }
            model_params = {**other_dict, **kwargs}
            # edge loop
            if edge_start is not None:
                if parallel:
                    Parallel(n_jobs=parallel_jobs_n)(
                        delayed(wrap_non_picklable_objects(edge_loop_for_models))(edge, **model_params)
                        for edge in range(kwargs["num_iter"])
                    )
                else:
                    for edge in range(edge_start, edge_end + 1):
                        start = time()
                        edge_loop_for_models(edge, **model_params)
                        elapsed_time = time() - start
                        print(f"Processing time is {round(elapsed_time, 3)}.")
        else:
            # calculate mean and/or pc scores of trait data
            trait_data = set_mean_pc_scores_of_trait(
                trait_type,
                kwargs["data_dict"]["trait_data"],
                model,
                scale_name,
                pca_z_scaling,
            )
            if edge_start is not None:
                for edge in range(edge_start, edge_end + 1):
                    edge_loop_for_nomodels(
                        kwargs["kfold_i"],
                        model,
                        model_index,
                        edge,
                        kwargs.get("memmap_dict").get("cor_memmap"),
                        kwargs["data_dict"],
                        **input_params,
                    )


def main():
    """
    main function for model estimations using sub-functions
    """
    print('Selecting data.')
    (
        fc_data,
        ffi_data,
        cog_data,
        asr_data,
        subjects_set_for_analysis,
        subjects_with_mri_folder,
    ) = select_data_for_analysis_mod(**input_params)

    subjects_n = len(subjects_set_for_analysis)
    subjects_list = list(subjects_set_for_analysis)
    
    family_df = pd.read_csv(FAMILY_PATH)
    family_df['Subject'] = family_df['Subject'].astype(str)
    family_df.query('Subject in @subjects_list', inplace=True)
    unique_families = family_df['Family_ID'].unique()
#    train_family_ids, test_family_ids = train_test_split(unique_family, test_size=1/k_fold, random_state=random_seed) 
    split_ratio = input_params["split_ratio"]
    
    if split_ratio is not None:
        kfold_num = int(1 / split_ratio)
        seed(random_seed)
        shuffle(unique_families)
    else:
        kfold_num, split_ratio = 1, 1

    if split_half_family:
        kfold_num = 2
    
    families_split_n = len(unique_families) // kfold_num
    #subjects_split_n = subjects_n // kfold_num

    def subjects_dict_process(kfold_i, kfold_num):
        """
        Conduct processing after splitting subjects
        """
        subjects_dict = {"split1": subjects_split1, "split2": subjects_split2}
        for subjects_set_key in subjects_dict.keys():

            if kfold_num == 2:
                if subjects_set_key == 'split1':
                    kfold_i = 0
                elif subjects_set_key == 'split2':
                    kfold_i = 1

            if not len(subjects_dict.get(subjects_set_key)) == 0:
                subjects_set_for_analysis = subjects_dict.get(subjects_set_key)
                (
                    data_dict,
                    memmap_dict,
                    logger_dict,
                    string_dict,
                    num_iter,
                    sample_n,
                    suffix,
                ) = settings_and_preprocesses(
                    subjects_set_for_analysis,
                    subjects_with_mri_folder,
                    fc_data,
                    ffi_data,
                    cog_data,
                    asr_data,
                    kfold_i,
                    **input_params,
                )
                settings_params = {
                    "data_dict": data_dict,
                    "string_dict": string_dict,
                    "memmap_dict": memmap_dict,
                    "logger_dict": logger_dict,
                    "num_iter": num_iter,
                    "sample_n": sample_n,
                    "kfold_i": kfold_i,
                    "suffix": suffix,
                }
                print('Start loop for estimation')
                model_loop_params = {**input_params, **settings_params}
                model_loop(**model_loop_params)
    
    output_subjects_dict = defaultdict(dict)
    for kfold_i in range(kfold_num):
        if kfold_num > 2:
            #subjects_split1, subjects_split2 = next(gkf.split(subjects_list, groups=None))
            if kfold_i + 1 < kfold_num:
                family_split_test = set(
                    unique_families[
                        kfold_i*families_split_n:(kfold_i + 1)*families_split_n
                    ]
                )
            else:
                family_split_test = set(unique_families[kfold_i*families_split_n:])
            family_split_train = {i for i in unique_families if i not in family_split_test}
            subjects_split1 = family_df.query('Family_ID in @family_split_train')['Subject'].to_list()
            output_subjects_dict[kfold_i] = defaultdict(dict)
            output_subjects_dict[kfold_i]['train'] = subjects_split1
            output_subjects_dict[kfold_i]['test'] = [i for i in subjects_list if i not in subjects_split1]
            subjects_split2 = []
            if run_main_process:
                subjects_dict_process(kfold_i=kfold_i, kfold_num=kfold_num)
#        elif kfold_num == 2 and not split_half_family:
#            if kfold_i == 0:
#                subjects_split1 = set(
#                        subjects_list[
#                            kfold_i * subjects_split_n : (kfold_i + 1) * subjects_split_n
#                        ]
#                    )
#                subjects_split2 = {i for i in subjects_list if i not in subjects_split1}
#            else:
#                break
        elif kfold_num == 2 and split_half_family:
            if kfold_i == 0:
                subjects_split1, subjects_split2 = train_test_split_family(random_seed=random_seed, k_fold=2, select_cb=select_cb)
                subjects_dict_process(kfold_i=kfold_i, kfold_num=kfold_num)
            else: 
                break
        elif kfold_num == 1:
            subjects_split1, subjects_split2 = subjects_list, []
            subjects_dict_process(kfold_i=kfold_i, kfold_num=kfold_num)
    
    if save_subjects_dict:
        print('Saving dictionary of subjects in CV')
        with open(f'/home/cezanne/t-haitani/hcp_data/derivatives/subjects_{kfold_num}folds_CV.pickle', 'wb') as f:
            pickle.dump(output_subjects_dict, f)

if __name__ == "__main__":
    # Give arguments from bash to python
    parser = argparse.ArgumentParser()
    
    parser.add_argument("--edge_start", type=int, default=None)
    parser.add_argument('--edge_end', type=int, default=None)
    parser.add_argument('--test', action='store_true')
    parser.add_argument('--array_id', type=int)
    parser.add_argument('--edge_total', type=int)
    parser.add_argument('--edge_divide_n', type=int)

    # FC data options
    parser.add_argument("--fc_filename", type=str)
    parser.add_argument("--exclude_subcortex", action="store_true")
    parser.add_argument("--fisher_z", action="store_true")
    parser.add_argument("--fc_unit", default='session')
    parser.add_argument('--invalid_edge_file', type=str)
    # Trait data options
    parser.add_argument("--trait_type", type=str)
    parser.add_argument("--scale_name", type=str)
    parser.add_argument('--remove_vars_list', nargs='+', type=str, default=None)
    # Saving options
    parser.add_argument("--save_data", action="store_true")
    parser.add_argument("--save_subjects_dict", action='store_true')
    parser.add_argument("--run_main_process", action='store_true')
    parser.add_argument("--dtype_memmap", type=str, default="float32")
    parser.add_argument("--save_cfa_data", action="store_true")
    # Model options
    parser.add_argument(
        "--model_type",
        nargs="*",
        default=["full", "mean", "pca", "model_fc", "model_trait", "model_both"],
    )
    parser.add_argument("--cov_cor", action="store_true")
    parser.add_argument("--phase_encoding", action="store_true")
    parser.add_argument("--day_cor", action="store_true")
    parser.add_argument("--order_in_day", action="store_true")
    parser.add_argument("--m_minus_1", action='store_true')
    parser.add_argument("--single_trait", action='store_true')
    parser.add_argument("--multistate_single_trait", action='store_true')
    parser.add_argument("--bi_factor", action='store_true')
    parser.add_argument("--diff_load", action='store_true')
    parser.add_argument("--CU", action="store_true")
    parser.add_argument("--OM", action='store_true')
    parser.add_argument('--std_lv', action='store_true')
    parser.add_argument('--trait_equal_loadings', action='store_true')
    parser.add_argument('--add_trait_error', action='store_true')
    parser.add_argument('--get_std_solutions', action='store_true')
    parser.add_argument("--control", nargs="*", type=str, default=["age", "gender"])
    parser.add_argument("--controlBefore", nargs="*", type=str, default=None)
    parser.add_argument("--model_fit_obj", default="MLR")
    parser.add_argument("--ordered", action="store_true")
    parser.add_argument("--cor_type", default="pearson", type=str)
    parser.add_argument("--se_robust", action="store_true")
    parser.add_argument("--pca_z_scaling", action="store_true")
    parser.add_argument("--fix_loadings_to_one", action="store_true")
    parser.add_argument("--fix_cov_to_zero", action="store_true")
    parser.add_argument("--fc_to_trait", action='store_true')
    parser.add_argument("--save_residuals", action="store_true")
    parser.add_argument("--use_lavaan", action="store_true")
    parser.add_argument("--cfa_standardize", action="store_true")
    parser.add_argument("--add_method_marker", action="store_true")
    parser.add_argument('--method_cor', action='store_true')
    parser.add_argument('--mean_structure', action='store_true')
    # Model convergence options
    parser.add_argument('--rel_tol', type=float, default=1e-10)
    parser.add_argument('--iter_max', type=int, default=10000)
    # Split options
    parser.add_argument("--split_half_family", action='store_true')
    parser.add_argument("--split_ratio", type=float, default=None)
    parser.add_argument("--random_seed", type=int, default=0)
    # Paralle processing option
    parser.add_argument("--parallel", action="store_true")
    parser.add_argument("--parallel_jobs_n", type=int)
    # Subject selection options
    parser.add_argument("--fmri_run_all_or_any", type=str, default="all")
    parser.add_argument("--trait_all_or_any", type=str, default="all")
    parser.add_argument("--cog_missing", type=str, default="remove")
    parser.add_argument("--rms_removal", action="store_true")
    parser.add_argument("--rms_remove_percentage", type=float)
    parser.add_argument("--rms_thres", type=float)
    parser.add_argument("--rms_pass_all_or_any", type=str)
    parser.add_argument("--select_cb", action='store_true')

    inputs = parser.parse_args()

    # Process inputs from shell
    print(inputs, flush=True)
    
    run_main_process = inputs.run_main_process

    edge_start = inputs.edge_start if not inputs.edge_start == -999 else None
    edge_end = inputs.edge_end if not inputs.edge_end == -999 else None
    num_iter = edge_end - edge_start + 1
    array_id = inputs.array_id
    edge_total = inputs.edge_total
    edge_divide_n = inputs.edge_divide_n

    test = inputs.test

    # FC data inputs
    fc_filename = inputs.fc_filename
    fisher_z = inputs.fisher_z
    exclude_subcortex = inputs.exclude_subcortex
    fc_unit = inputs.fc_unit
    invalid_edge_file = inputs.invalid_edge_file

    if 'Schaefer' in fc_filename:
        parcellation = 'Schaefer'
    elif 'Gordon' in fc_filename:
        parcellation = 'Gordon'
    ATLAS_DIR = ATLAS_DIR_DICT.get(parcellation)
    FC_ONLY_DIR = FC_ONLY_DIR_DICT.get(parcellation)
    NIH_COGNITION_DIR = NIH_COGNITION_DIR_DICT.get(parcellation)
    ASR_DIR = ASR_DIR_DICT.get(parcellation)
    NEO_FFI_DIR = NEO_FFI_DIR_DICT.get(parcellation)

    # Trait data inputs
    trait_type = inputs.trait_type
    scale_name = inputs.scale_name
    remove_vars_list = inputs.remove_vars_list
    if remove_vars_list:
        remove_vars_list = remove_vars_list if not inputs.remove_vars_list[0] in ['None,', 'None'] else None
    
    # Saving option inputs
    save_data = inputs.save_data
    save_subjects_dict = inputs.save_subjects_dict
    save_cfa_data = inputs.save_cfa_data
    dtype_memmap = inputs.dtype_memmap
    
    # Modeling inputs
    model_type = inputs.model_type
    cor_type = inputs.cor_type
    pca_z_scaling = inputs.pca_z_scaling
    control = None if "".join(inputs.control) == "None" else inputs.control
    control_before = inputs.controlBefore
    equal_loadings = inputs.trait_equal_loadings
    #control_before = None if "".join(inputs.controlBefore) == "None" else inputs.controlBefore
    ordered = inputs.ordered
    fix_loadings_to_one = inputs.fix_loadings_to_one
    fix_cov_to_zero = inputs.fix_cov_to_zero
    day_cor = inputs.day_cor
    fc_to_trait = inputs.fc_to_trait
    model_fit_obj = inputs.model_fit_obj
    se_robust = inputs.se_robust
    cov_cor = inputs.cov_cor
    CU = inputs.CU
    OM = inputs.OM
    phase_encoding = inputs.phase_encoding
    order_in_day = inputs.order_in_day
    m_minus_1 = inputs.m_minus_1
    single_trait = inputs.single_trait
    multistate_single_trait = inputs.multistate_single_trait
    bi_factor = inputs.bi_factor
    diff_load = inputs.diff_load
    std_lv = inputs.std_lv
    get_std_solutions = inputs.get_std_solutions
    save_residuals = inputs.save_residuals
    use_lavaan = inputs.use_lavaan
    cfa_standardize = inputs.cfa_standardize
    add_method_marker = inputs.add_method_marker
    method_cor = inputs.method_cor
    mean_structure = inputs.mean_structure
    add_trait_error = inputs.add_trait_error

    # Model convergence inputs
    iter_max = inputs.iter_max
    rel_tol = inputs.rel_tol

    # Splitting option inputs
    split_half_family = inputs.split_half_family
    split_ratio = inputs.split_ratio
    random_seed = inputs.random_seed
    print(type(random_seed))
    # Parallel option input
    parallel = inputs.parallel
    parallel_jobs_n = inputs.parallel_jobs_n
    
    # Subject selection option inputs
    rms_removal = inputs.rms_removal
    rms_remove_percentage = inputs.rms_remove_percentage
    rms_thres = inputs.rms_thres
    rms_pass_all_or_any = inputs.rms_pass_all_or_any
    fmri_run_all_or_any = inputs.fmri_run_all_or_any
    trait_all_or_any = inputs.trait_all_or_any
    cog_missing = inputs.cog_missing
    select_cb = inputs.select_cb
    
    input_params = {
        'edge_start': edge_start,
        'edge_end': edge_end,
        # FC
        "fc_filename": fc_filename,
        "fisher_z": fisher_z,
        "exclude_subcortex": exclude_subcortex,
        "fc_unit": fc_unit,
        'invalid_edge_file': invalid_edge_file,
        "parcellation": parcellation,
        # Trait
        "trait_type": trait_type,
        "scale_name": scale_name,
        "remove_vars_list": remove_vars_list,
        # Model
        "model_type": model_type,
        "control": control,
        "day_cor": day_cor,
        "se_robust": se_robust,
        "model_fit_obj": model_fit_obj,
        "cor_type": cor_type,
        "ordered": ordered,
        "fix_loadings_to_one": fix_loadings_to_one,
        'equal_loadings': equal_loadings,
        "add_trait_error": add_trait_error,
        'fix_cov_to_zero': fix_cov_to_zero,
        "pca_z_scaling": pca_z_scaling,
        "cov_cor": cov_cor,
        "OM": OM,
        "phase_encoding": phase_encoding,
        'order_in_day': order_in_day,
        "m_minus_1": m_minus_1,
        "single_trait": single_trait,
        "multistate_single_trait": multistate_single_trait,
        "bi_factor": bi_factor,
        "diff_load": diff_load,
        "std_lv": std_lv,
        "get_std_solutions": get_std_solutions,
        "fc_to_trait": fc_to_trait,
        "use_lavaan": use_lavaan,
        "cfa_standardize": cfa_standardize,
        "mean_structure": mean_structure,
        # Saving
        "save_data": save_data,
        "save_cfa_data": save_cfa_data,
        "dtype_memmap": dtype_memmap,
        "save_residuals": save_residuals,
        # Splitting
        "split_half_family": split_half_family,
        "split_ratio": split_ratio,
        # Parallel processing
        "parallel": parallel,
        "parallel_jobs_n": parallel_jobs_n,
        # Subject selection
        "cog_missing": cog_missing,
        "fmri_run_all_or_any": fmri_run_all_or_any,
        "trait_all_or_any": trait_all_or_any,
        "rms_removal": rms_removal,
        "rms_remove_percentage": rms_remove_percentage,
        "rms_thres": rms_thres,
        "rms_pass_all_or_any": rms_pass_all_or_any,
        "select_cb": select_cb
    }
    trait_scale_name = get_scale_name_from_trait(trait_type)
    main()
